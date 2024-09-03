#=============================================================================
# Source Modules
#=============================================================================
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
import sys
try:
    from classes import pcolor, validating
    from copy import deepcopy
    from Crypto.PublicKey import RSA
    from datetime import datetime, timedelta
    from dotmap import DotMap
    from ecdsa import SigningKey
    from git import cmd, Repo
    from json_ref_dict import materialize, RefDict
    from openpyxl import load_workbook
    from OpenSSL import crypto
    from pathlib import Path
    from stringcase import snakecase
    import argparse, base64, crypt, ipaddress, itertools, jinja2, json, logging, os, pexpect, pkg_resources, platform
    import pytz, re, requests, shutil, subprocess, stdiomask, string, textwrap, validators, yaml
except ImportError as e:
    prRed(f'!!! ERROR !!!\n{e.__class__.__name__}')
    prRed(f' Module {e.name} is required to run this script')
    prRed(f' Install the module using the following: `pip install {e.name}`')
    sys.exit(1)
#=============================================================================
# Log levels 0 = None, 1 = Class only, 2 = Line
#=============================================================================
log_level = 2
#=============================================================================
# Exception Classes and YAML dumper
#=============================================================================
class insufficient_args(Exception): pass

class yaml_dumper(yaml.Dumper):
    def increase_indent(self, flow=False, indentless=False):
        return super(yaml_dumper, self).increase_indent(flow, False)

#=============================================================================
# Function - Arguments used by the majority of the modules
#=============================================================================
def base_arguments(parser):
    parser.add_argument(
        '-a', '--intersight-api-key-id', default = os.getenv('intersight_api_key_id'),
        help   = 'The Intersight API key id for HTTP signature scheme.')
    parser.add_argument(
        '-d', '--dir', default = 'Intersight',
        help   = 'The Directory to use for the Creation of the YAML Configuration Files.')
    parser.add_argument(
        '-dl', '--debug-level', default = 0,
        help ='Used for troubleshooting.  The Amount of Debug output to Show: '\
            '1. Shows the api request response status code '\
            '5. Show URL String + Lower Options '\
            '6. Adds Results + Lower Options '\
            '7. Adds json payload + Lower Options '\
            'Note: payload shows as pretty and straight to check for stray object types like Dotmap and numpy')
    parser.add_argument(
        '-f', '--intersight-fqdn', default ='intersight.com',
        help   = 'The Directory to use for the Creation of the YAML Configuration Files.')
    parser.add_argument(
        '-i', '--ignore-tls', action = 'store_false',
        help   = 'Ignore TLS server-side certificate verification.  Default is False.')
    parser.add_argument('-j', '--json-file', default = None, help   = 'The IMM Transition Tool JSON Dump File to Convert to HCL.')
    parser.add_argument(
        '-k', '--intersight-secret-key', default = os.getenv('intersight_secret_key'),
        help   = 'Name of the file containing The Intersight secret key or contents of the secret key in environment.')
    parser.add_argument('-rc', '--repository-check-skip', action = 'store_true', help   = 'Flag to Skip Repository URL Test for OS Install.')
    parser.add_argument('-l',  '--load-config',      action = 'store_true', help   = 'Skip Wizard and Just Load Configuration Files.')
    parser.add_argument('-y', '--yaml-file', default = None,  help = 'The input YAML File.')
    return parser

#=============================================================================
# Function - Arguments used by EZIMM For Sensitive Variables
#=============================================================================
def base_arguments_ezimm_sensitive_variables(parser):
    parser.add_argument('-alp',  '--azure-stack-lcm-password',      help='Azure Stack HCI Life Cycle Management User Password.')
    parser.add_argument('-ccp',  '--cco-password',                  help='CCO Password to Authorize Firmware Downloads.')
    parser.add_argument('-ccu',  '--cco-user',                      help='CCO Username to Authorize Firmware Downloads.')
    parser.add_argument('-dap',  '--domain-administrator-password', help='Windows Domain Administrator Password.')
    parser.add_argument('-ilp',  '--local-user-password-1',         help='Intersight Managed Mode Local User Password 1.')
    parser.add_argument('-ilp2', '--local-user-password-2',         help='Intersight Managed Mode Local User Password 2.')
    parser.add_argument('-imm',  '--imm-transition-password',       help='IMM Transition Tool Password.')
    parser.add_argument('-isa',  '--snmp-auth-password-1',          help='Intersight Managed Mode SNMP Auth Password.')
    parser.add_argument('-isp',  '--snmp-privacy-password-1',       help='Intersight Managed Mode SNMP Privilege Password.')
    parser.add_argument('-lap',  '--local-administrator-password',  help='Windows Local Administrator Password.')
    parser.add_argument('-np',   '--netapp-password',               help='NetApp Login Password.')
    parser.add_argument('-nsa',  '--netapp-snmp-auth',              help='NetApp SNMP Auth Password.')
    parser.add_argument('-nsp',  '--netapp-snmp-priv',              help='NetApp SNMP Privilege Password.')
    parser.add_argument('-nxp',  '--nexus-password',                help='Nexus Login Password.')
    parser.add_argument('-p',    '--pure-storage-password',         help='Pure Storage Login Password.')
    parser.add_argument('-psa',  '--pure-storage-snmp-auth',        help='Pure Storage SNMP Auth Password.')
    parser.add_argument('-psp',  '--pure-storage-snmp-priv',        help='Pure Storage SNMP Privilege Password.')
    parser.add_argument('-pxp',  '--proxy-password',                help='Proxy Password.')
    parser.add_argument('-vep',  '--vmware-esxi-password',          help='VMware ESXi Root Login Password.')
    parser.add_argument('-vvp',  '--vmware-vcenter-password',       help='VMware vCenter Admin Login Password.')
    return parser

#=============================================================================
# Function - Basic Setup for the Majority of the modules
#=============================================================================
def base_script_settings(kwargs):
    #=========================================================================
    # Configure logger and Build kwargs
    #=========================================================================
    script_name = (sys.argv[0].split(os.sep)[-1]).split('.')[0]
    dest_dir    = f'{Path.home()}{os.sep}Logs'
    dest_file   = script_name + '.log'
    if not os.path.exists(dest_dir): os.mkdir(dest_dir)
    if not os.path.exists(os.path.join(dest_dir, dest_file)): 
        create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'; os.system(create_file)
    FORMAT = '%(asctime)-15s [%(levelname)s] [%(filename)s:%(lineno)s] %(message)s'
    logging.basicConfig(filename=f'{dest_dir}{os.sep}{script_name}.log', filemode='a', format=FORMAT, level=logging.DEBUG )
    logger = logging.getLogger('openapi')
    #=========================================================================
    # Determine the Script Path
    #=========================================================================
    args_dict = vars(kwargs.args)
    for k,v in args_dict.items():
        if type(v) == str and v != None: os.environ[k] = v
    kwargs.script_name   = (sys.argv[0].split(os.sep)[-1]).split('.')[0]
    kwargs.script_path   = os.path.dirname(os.path.realpath(sys.argv[0]))
    kwargs.args.dir      = os.path.abspath(kwargs.args.dir)
    kwargs.home          = Path.home()
    kwargs.logger        = logger
    kwargs.op_system     = platform.system()
    kwargs.imm_dict.orgs = DotMap()
    kwargs.type_dotmap   = type(DotMap())
    kwargs.type_none     = type(None)
    #=========================================================================
    # Import Stored Parameters and Add to kwargs
    #=========================================================================
    ezdata          = materialize(RefDict(os.path.join(kwargs.script_path, 'variables', 'easy-imm.json'), 'r', encoding='utf8'))
    script_tag      = script_name.replace('ez', 'easy-')
    kwargs.ez_tags  = [{'Key':'Module','Value':script_tag},{'Key':'Version','Value':ezdata['info']['version']}]
    kwargs.ezdata   = DotMap(ezdata['components']['schemas'])
    kwargs.ezwizard = DotMap(ezdata['components']['wizard'])
    #=========================================================================
    # Get Intersight Configuration
    # - apikey
    # - endpoint
    # - keyfile
    #=========================================================================
    if kwargs.args.intersight_secret_key:
        if '~' in kwargs.args.intersight_secret_key: kwargs.args.intersight_secret_key = os.path.expanduser(kwargs.args.intersight_secret_key)
    if os.getenv('intersight_fqdn'): kwargs.args.intersight_fqdn = os.getenv('intersight_fqdn')
    if 'api/v1/iam/Users' in kwargs.args.intersight_fqdn: kwargs.args.intersight_fqdn = ((kwargs.args.intersight_fqdn).replace('https://', '')).split('/')[0]
    elif os.getenv('intersight_fqdn'): kwargs.args.intersight_fqdn = os.getenv('intersight_fqdn')
    kwargs          = intersight_config(kwargs)
    kwargs.args.url = 'https://%s' % (kwargs.args.intersight_fqdn)
    #=========================================================================
    # Check Folder Structure for Illegal Characters
    #=========================================================================
    for folder in kwargs.args.dir.split(os.sep):
        if folder == '': pass
        elif not re.search(r'^[\w\@\-\.\:\/\\]+$', folder):
            pcolor.Red(f'\n{"-"*108}\n\n  !!ERROR!!')
            pcolor.Red(f'  The Directory structure can only contain the following characters:')
            pcolor.Red(f'  letters(a-z, A-Z), numbers(0-9), hyphen(-), period(.), colon(:), and underscore(-).')
            pcolor.Red(f'  It can be a short path or a fully qualified path.  `{folder}` does not qualify.')
            pcolor.Red(f'  Exiting...\n\n{"-"*108}\n')
            len(False); sys.exit(1)
    return kwargs

#=============================================================================
# pexpect - Login Function
#=============================================================================
def child_login(kwargs):
    kwargs.sensitive_var = kwargs.password
    kwargs   = sensitive_var_value(kwargs)
    password = kwargs.var_value
    kwargs.password = password
    log_dir = os.path.join(str(Path.home()), 'Logs')
    if not os.path.isdir(log_dir): os.mkdir(log_dir)
    #=========================================================================
    # Launch Local Shell
    #=========================================================================
    if kwargs.op_system == 'Windows':
        from pexpect import popen_spawn
        child = popen_spawn.PopenSpawn('cmd', encoding='utf-8', timeout=1)
    else:
        system_shell = os.environ['SHELL']
        child = pexpect.spawn(system_shell, encoding='utf-8')
    child.logfile_read = sys.stdout
    #=========================================================================
    # Test Connectivity with Ping and then Login
    #=========================================================================
    if kwargs.op_system == 'Windows':
        child.sendline(f'ping -n 2 {kwargs.hostname}')
        child.expect(f'ping -n 2 {kwargs.hostname}')
        child.expect_exact('> ')
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | Tee-Object {log_dir}\{kwargs.hostname}.txt')
        child.expect(f'Tee-Object {log_dir}\{kwargs.hostname}.txt')
    else:
        child.sendline(f'ping -c 2 {kwargs.hostname}')
        child.expect(f'ping -c 2 {kwargs.hostname}')
        child.expect_exact('$ ')
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | tee {log_dir}/{kwargs.hostname}.txt')
        child.expect(f'tee {log_dir}/{kwargs.hostname}.txt')
    logged_in = False
    while logged_in == False:
        i = child.expect(['Are you sure you want to continue', 'closed', 'password:', 'Password:', kwargs.host_prompt, pexpect.TIMEOUT])
        if i == 0: child.sendline('yes')
        elif i == 1:
            pcolor.Red(f'\n!!! FAILED !!! to connect.  '\
                f'Please Validate {kwargs.hostname} is correct and username {kwargs.username} is correct.')
            sys.exit(1)
        elif i == 2: child.sendline(password)
        elif i == 3: child.sendline(password)
        elif i == 4: logged_in = True
        elif i == 5:
            pcolor.Red(f'\n{"-"*91}\n')
            pcolor.Red(f'!!! FAILED !!!\n Could not open SSH Connection to {kwargs.hostname}')
            pcolor.Red(f'\n{"-"*91}\n')
            sys.exit(1)
    # Return values
    return child, kwargs

#=============================================================================
# Function - Format Policy Description
#=============================================================================
def choose_policy(policy_type, kwargs):
    policy_descr = mod_pol_description(policy_type)
    policy_list = []
    for i in kwargs['policies'][policy_type]: policy_list.append(i['name'])
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n{"-"*108}\n')
        if kwargs.get('optional_message'): pcolor.Cyan(kwargs['optional_message'])
        pcolor.Cyan(f'  {policy_descr} Policy Options:')
        for i, v in enumerate(policy_list):
            i += 1
            if i < 10: pcolor.Cyan(f'     {i}. {v}')
            else: pcolor.Cyan(f'    {i}. {v}')
        if kwargs['allow_opt_out'] == True: pcolor.Cyan(f'     99. Do not assign a(n) {policy_descr}.')
        pcolor.Cyan(f'     100. Create a New {policy_descr}.')
        pcolor.Cyan(f'\n{"-"*108}\n')
        policyOption = input(f"Select the Option Number for the {policy_descr} Policy to Assign to {kwargs['name']} Policy: ")
        if re.search(r'^[0-9]{1,3}$', policyOption):
            for i, v in enumerate(policy_list):
                i += 1
                if   int(policyOption) == i:   kwargs['policy'] = v;  valid = True; return kwargs
                elif int(policyOption) == 99:  kwargs['policy'] = ''; valid = True; return kwargs
                elif int(policyOption) == 100: kwargs['policy'] = 'create_policy'; valid = True; return kwargs
            if   int(policyOption) == 99:  kwargs['policy'] = ''; valid = True; return kwargs
            elif int(policyOption) == 100: kwargs['policy'] = 'create_policy'; valid = True; return kwargs
        else: message_invalid_selection()

#=============================================================================
# Function - Count the Number of Keys
#=============================================================================
def count_keys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func: count += 1
    return count

#=============================================================================
# Function for Processing easyDict and Creating YAML Files
#=============================================================================
def create_yaml(orgs, kwargs):
    ezdata  = kwargs.ezdata.ezimm_class.properties
    classes = kwargs.ezdata.ezimm_class.properties.classes.enum
    def write_file(dest_dir, dest_file, idict, title):
        if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
        if not os.path.exists(os.path.join(dest_dir, dest_file)):
            create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
            os.system(create_file)
        wr_file = open(os.path.join(dest_dir, dest_file), 'w')
        wr_file.write('---\n')
        wr_file = open(os.path.join(dest_dir, dest_file), 'a')
        dash_length = '='*(len(title) + 20)
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(f'#   {title} - Variables\n')
        wr_file.write(f'#{dash_length}\n')
        if 'name_pfx_sfx' in dest_file: wr_file.write('# If a prefix/suffix policy is undefined default will be used\n')
        wr_file.write(yaml.dump(idict, Dumper = yaml_dumper, default_flow_style=False))
        wr_file.close()
    for item in classes:
        dest_dir = os.path.join(kwargs.args.dir, ezdata[item].directory)
        if re.search('policies|pools|profiles|templates', item):
            for i in ezdata[item].enum:
                idict = {}
                for org in orgs:
                    org_keys = list(kwargs.imm_dict.orgs[org].keys())
                    if item in org_keys:
                        if not idict.get(org): idict[org] = {}
                        if not idict[org].get(item): idict[org][item] = {}
                        ikeys = ezdata[i].enum
                        idict[org][item] = dict(sorted(deepcopy(kwargs.imm_dict.orgs[org][item].toDict()).items()))
                        pkeys = list(kwargs.imm_dict.orgs[org][item].keys())
                        for x in pkeys:
                            if not x in ikeys: idict[org][item].pop(x)
                            elif not len(idict[org][item][x]) > 0: idict[org][item].pop(x)
                        if len(idict[org][item]) == 0: idict.pop(org)
                if len(idict) > 0:
                    title     = mod_pol_description(f"{str.title(item.replace('_', ' '))} -> {str.title((ezdata[i].title).replace('_', ' '))}")
                    dest_file = f'{ezdata[i].title}.ezi.yaml'
                    write_file(dest_dir, dest_file, idict, title)
        else:
            for i in ezdata[item].enum:
                idict = {}
                if item == i:
                    for org in orgs:
                        org_keys = list(kwargs.imm_dict.orgs[org].keys())
                        if item in org_keys and len(kwargs.imm_dict.orgs[org][item]) > 0:
                            idict[org] = {}; idict[org][item] = dict(sorted(deepcopy(kwargs.imm_dict.orgs[org][item].toDict()).items()))
                else:
                    for org in orgs:
                        org_keys = list(kwargs.imm_dict.orgs[org].keys())
                        if item in org_keys:
                            if not idict.get(org): idict[org] = {}
                            if not idict[org].get(item): idict[org][item] = {}
                            ikeys = ezdata[i].enum
                            pkeys = list(kwargs.imm_dict.orgs[org][item].keys())
                            idict[org][item] = dict(sorted(deepcopy(kwargs.imm_dict.orgs[org][item].toDict()).items()))
                            for x in pkeys:
                                if not i == x or len(kwargs.imm_dict.orgs[org][item][i]) == 0: idict[org][item].pop(x)
                            if len(idict[org][item]) == 0: idict.pop(org)
                if len(idict) > 0:
                    if i == item:
                        title   = mod_pol_description(str.title(item.replace('_', ' ')))
                    else: title = mod_pol_description(f"{str.title(item.replace('_', ' '))} -> {str.title((i).replace('_', ' '))}")

                    write_file(dest_dir, f'{i}.yaml', idict, title)

#=============================================================================
# Function - Cleanup Empty Parameters in Dictionary
#=============================================================================
def dictionary_cleanup(dictionary):
    none_type = type(None)
    for k in list(dictionary.keys()):
        if type(dictionary[k]) == none_type: dictionary.pop(k)
        elif type(dictionary[k]) == str  and dictionary[k] == '':     dictionary.pop(k)
        elif type(dictionary[k]) == list and len(dictionary[k]) == 0: dictionary.pop(k)
    return dictionary

#=============================================================================
# Function - Prompt User with question
#=============================================================================
def exit_confirm_loop(kwargs):
    question = 'Y'
    valid_confirm = False
    while valid_confirm == False:
        question = input(f'Do you want to accept the above configuration?  Enter "Y" or "N" [{kwargs.yes_or_no}]: ')
        if question == '': question = kwargs.yes_or_no
        if question == 'Y':
            kwargs.accept_configuration = True
            valid_exit = False
            while valid_exit == False:
                loop_exit = input(
                    f'Would You like to Configure another {kwargs.policy_type}?  Enter "Y" or "N" [{kwargs.yes_or_no}]: ')
                if loop_exit == '': loop_exit = 'N'
                if loop_exit == 'Y':   valid_confirm = True; valid_exit = True; kwargs.configure_additional = True
                elif loop_exit == 'N': valid_confirm = True; valid_exit = True; kwargs.configure_additional = False
                else: message_invalid_y_or_n('short')
        elif question == 'N':
            kwargs.accept_configuration = False
            kwargs.configure_additional = True
            message_starting_over(kwargs.policy_type)
            valid_confirm = True
        else: message_invalid_y_or_n('long')
    return kwargs

#=============================================================================
# Function - Ask User to Configure Additional Policy
#=============================================================================
def exit_default(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '':    exit_answer = y_or_n
        if exit_answer == 'N':   configure_loop = True;  policy_loop = True;  valid_exit = True
        elif exit_answer == 'Y': configure_loop = False; policy_loop = False; valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#=============================================================================
# Function - Ask User to Configure Additional Policy
#=============================================================================
def exit_default_del_tfc(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '':    exit_answer = y_or_n
        if exit_answer == 'N':   policy_loop = True;  configure_loop = True;  valid_exit = True
        elif exit_answer == 'Y': policy_loop = False; configure_loop = False; valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#=============================================================================
# Function - Prompt User with question
#=============================================================================
def exit_loop_default_yes(loop_count, policy_type):
    valid_exit = False
    while valid_exit == False:
        if loop_count % 2 == 0:
            exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [Y]: ')
        else: exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [N]: ')
        if (loop_count % 2 == 0 and exit_answer == '') or exit_answer == 'Y':
            configure_loop = False; loop_count += 1; policy_loop = False; valid_exit = True
        elif not loop_count % 2 == 0 and exit_answer == '':
            configure_loop = True;  loop_count += 1; policy_loop = True;  valid_exit = True
        elif exit_answer == 'N':
            configure_loop = True;  loop_count += 1; policy_loop = True;  valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, loop_count, policy_loop

#=============================================================================
# Function to Append the imm_dict Dictionary
#=============================================================================
def ez_append(pol_vars, kwargs):
    class_path= kwargs.class_path
    p         = class_path.split(',')
    if kwargs.use_shared_org == True and re.search('^policies|pools', class_path):
        org   = kwargs.shared_org
    else: org = kwargs.org
    pol_vars = DotMap(ez_remove_empty(pol_vars))
    # Confirm the Key Exists
    if not kwargs.imm_dict.orgs.get(org): kwargs.imm_dict.orgs[org] = DotMap()
    if len(p) >= 2:
        if not kwargs.imm_dict.orgs[org].get(p[0]): kwargs.imm_dict.orgs[org][p[0]] = DotMap()
    if len(p) >= 3:
        if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]): kwargs.imm_dict.orgs[org][p[0]][p[1]] = DotMap()
    if len(p) >= 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]): kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = DotMap()
    if len(p) == 2:
        if kwargs.append_type == 'map': kwargs.imm_dict.orgs[org][p[0]][p[1]] = DotMap(pol_vars.toDict())
        else:
            if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]): kwargs.imm_dict.orgs[org][p[0]][p[1]] = [deepcopy(pol_vars)]
            else: kwargs.imm_dict.orgs[org][p[0]][p[1]].append(deepcopy(pol_vars))
    elif len(p) == 3:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]): kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = [deepcopy(pol_vars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].append(deepcopy(pol_vars))
    elif len(p) == 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]): kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = [deepcopy(pol_vars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].append(deepcopy(pol_vars))
    elif len(p) == 5:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]): kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = DotMap()
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].get(p[4]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]] = [deepcopy(pol_vars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]].append(deepcopy(pol_vars))
    kwargs.append_type = 'list'
    return kwargs

#=============================================================================
# Function to Append the imm_dict Dictionary
#=============================================================================
def ez_append_wizard(pol_vars, kwargs):
    class_path = kwargs['class_path']
    p = class_path.split(',')
    pol_vars = ez_remove_empty(pol_vars)
    # Confirm the Key Exists
    if not kwargs.imm_dict.get('wizard'): kwargs.imm_dict.wizard = {}
    if len(p) >= 2:
        if not kwargs.imm_dict.wizard.get(p[0]): kwargs.imm_dict.wizard.update(deepcopy({p[0]:{}}))
    if len(p) >= 3:
        if not kwargs.imm_dict.wizard[p[0]].get(p[1]): kwargs.imm_dict.wizard[p[0]].update(deepcopy({p[1]:{}}))
    if len(p) == 1:
        if not kwargs.imm_dict.wizard.get(p[0]): kwargs.imm_dict.wizard.update(deepcopy({p[0]:[]}))
    elif len(p) == 2:
        if not kwargs.imm_dict.wizard[p[0]].get(p[1]): kwargs.imm_dict.wizard[p[0]].update(deepcopy({p[1]:[]}))
    elif len(p) == 3:
        if not kwargs.imm_dict.wizard[p[0]][p[1]].get(p[2]): kwargs.imm_dict.wizard[p[0]][p[1]].update(deepcopy({p[2]:[]}))
    # append the Dictionary
    if len(p) == 1: kwargs.imm_dict.wizard[p[0]].append(deepcopy(pol_vars))
    if len(p) == 2: kwargs.imm_dict.wizard[p[0]][p[1]].append(deepcopy(pol_vars))
    elif len(p) == 3: kwargs.imm_dict.wizard[p[0]][p[1]][p[2]].append(deepcopy(pol_vars))
    return kwargs

#=============================================================================
# Function to Remove Empty Arguments
#=============================================================================
def ez_remove_empty(pol_vars):
    pop_list = []
    for k,v in pol_vars.items():
        if v == None: pop_list.append(k)
    for i in pop_list: pol_vars.pop(i)
    return pol_vars

#=============================================================================
# Function - find the Keys for each Section
#=============================================================================
def find_keys(ws, func_regex):
    func_list = {}
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)): func_list.add(str(i[0].value))
    func_list = DotMap(dict(sorted(func_list.items())))
    return func_list

#=============================================================================
# Function - Assign the Variables to the Keys
#=============================================================================
def find_vars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value: var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else: x += 1
            except Exception as e: e = e; pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list: var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value; var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#=============================================================================
# Function - Build api_body for Operating System Installation - VMware
#=============================================================================
def installation_body(v, kwargs):
    if kwargs.script_name == 'ezci' and kwargs.args.deployment_type == 'azure_stack':
        api_body = installation_body_azure_stack(v, kwargs)
    elif kwargs.script_name == 'ezci' and kwargs.args.deployment_type == 'vmware':
        api_body = installation_body_vmware(v, kwargs)
    elif 'shared' in kwargs.os_cfg_moids[v.os_configuration].Owners:
        answers = DotMap(); encrypted = False
        answer_keys = list(v.answers.keys())
        for e in kwargs.os_cfg_moids[v.os_configuration].Placeholders:
            if re.search('\.answers\.', e.Type.Name): name = e.Type.Name[9:]
            elif '.internal' in e.Type.Name: continue
            elif 'FQDN' in e.Type.Name: continue
            else: name = e.Type.Name[1:]
            x = name.split('.')
            if x[0] in answer_keys:
                if type(v.answers[x[0]]) == str and 'sensitive_' in v.answers[x[0]]:
                    kwargs.sensitive_var = v.answers[x[0]].replace('sensitive_', '')
                    kwargs   = sensitive_var_value(kwargs)
                    password = kwargs.var_value
                    if v.os_vendor == 'Microsoft':
                        if 'LogonPassword' in x[0]:
                            answers[x[0]]   = base64.b64encode(f'{password}Password'.encode(encoding='utf-16-le')).decode()
                        else: answers[x[0]] = base64.b64encode(f'{password}AdministratorPassword'.encode(encoding='utf-16-le')).decode()
                    else: answers[x[0]] = crypt.crypt(password, crypt.mksalt(crypt.METHOD_SHA512))
                    encrypted = True
                elif x[0] == 'NameServer': answers['Nameserver'] = v.answers[x[0]]
                elif x[0] == 'AlternateNameServer': answers['AlternateNameServers'] = [v.answers['AlternateNameServer']]
                else: answers[x[0]] = v.answers[x[0]]
        if answers.get('IpV4Config') or answers.get('IpV6Config') or answers.IpConfigType == 'DHCP':
            vx = answers.IpVersion
            answers.pop('IpVersion')
            if answers.IpConfigType == 'DHCP': ip_config = {'ObjectType': f'os.Ip{vx.lower()}Configuration'}
            else: ip_config = DotMap({f'Ip{vx}Config': answers[f'Ip{vx}Config'].toDict(), 'ObjectType': f'os.Ip{vx.lower()}Configuration'}).toDict()
            answers.IpConfiguration = ip_config
            if f'Ip{vx}Config' in answer_keys: answers.pop(f'Ip{vx}Config')
        if answers.get('FQDN') and v.os_vendor == 'VMware': answers.Hostname = answers.FQDN; answers.pop('FQDN')
        api_body = {
            'Answers': dict(sorted(dict(answers, **{'IsRootPasswordCrypted': False, 'Source': 'Template'}).items())),
            'AdditionalParameters': None,
            'Description': '',
            'ConfigurationFile': {'Moid': v.os_configuration, 'ObjectType': 'os.ConfigurationFile'},
            'Image': {'Moid': v.os_image, 'ObjectType': 'softwarerepository.OperatingSystemFile'},
            'InstallMethod': 'vMedia',
            'InstallTarget': {},
            'ObjectType': 'os.Install', 
            'OperatingSystemParameters': {'Edition': 'Standard','ObjectType': 'os.WindowsParameters'},
            'Organization': {'Moid': kwargs.org_moids[kwargs.org].moid, 'ObjectType': 'organization.Organization'},
            'OsduImage': {'Moid': v.scu, 'ObjectType': 'firmware.ServerConfigurationUtilityDistributable'},
            'OverrideSecureBoot': False,
            'Server': {'Moid': v.hardware_moid, 'ObjectType': v.object_type}}
        if v.os_vendor == 'Microsoft':
            api_body['OperatingSystemParameters']['Edition'] = v.answers.EditionString
            api_body['Answers'].pop('EditionString')
        else: api_body.pop('OperatingSystemParameters')
        if api_body['Answers'].get('SecureBoot'):
            api_body['Answers'].pop('SecureBoot')
            if v.boot_order.enable_secure_boot == True: api_body.update({'OverrideSecureBoot': True})
        if api_body['Answers'].get('BootMode'): api_body['Answers'].pop('BootMode')
        if encrypted == True: api_body['Answers']['IsRootPasswordCrypted'] = True
    if v.boot_volume.lower() == 'm2':
        for k,v in v.storage_controllers.items():
            vkeys = list(v.keys())
            slot_rx = re.compile('(MSTOR-RAID)')
            if 'virtual_drives' in vkeys and re.search(slot_rx, v.slot):
                drive = [DotMap(id = a, name = b.name, slot = re.search(slot_rx, v.slot).group(1)) for a,b in v.virtual_drives.items() if int(b.size) > 128000][0]
                break
        api_body['InstallTarget'] = {'Id': str(drive.id), 'Name': drive.name, 'ObjectType': 'os.VirtualDrive', 'StorageControllerSlotId': drive.slot}
    elif v.boot_volume.lower() == 'san':
        api_body['InstallTarget'] = {'InitiatorWwpn': kwargs.fc_ifs[kwargs.san_target.interface_name].wwpn, 'LunId': kwargs.san_target.lun,
                                     'ObjectType': 'os.FibreChannelTarget', 'TargetWwpn': kwargs.san_target.wwpn}
    api_body = dict(sorted(api_body.items()))
    return api_body

#=============================================================================
# Function - Build api_body for Operating System Installation - Azure Stack
#=============================================================================
def installation_body_azure_stack(v, kwargs):
    vnics = []
    for a,b in v.adapters.items():
        bkeys = list(b.keys())
        if 'eth_ifs' in bkeys: vnics.append(b.adapter_id)
    vnics.sort()
    api_body = {
        'AdditionalParameters': [],
        'Answers': {'Source': 'Template'},
        'ConfigurationFile': {'Moid': v.os_configuration, 'ObjectType': 'os.ConfigurationFile'},
        'Description': '',
        'Image': {'Moid': v.os_image, 'ObjectType': 'softwarerepository.OperatingSystemFile'},
        'InstallMethod': 'vMedia',
        'OperatingSystemParameters': {'Edition': 'DatacenterCore', 'ObjectType': 'os.WindowsParameters'},
        'Organization': {'Moid': kwargs.org_moids[kwargs.org].moid, 'ObjectType': 'organization.Organization'},
        'OsduImage': {'Moid': v.scu, 'ObjectType': 'firmware.ServerConfigurationUtilityDistributable'},
        'OverrideSecureBoot': True,
        'Server': {'Moid': v.hardware_moid, 'ObjectType': v.object_type}}
    ad = kwargs.imm_dict.wizard.azure_stack[0].active_directory
    answers_dict = {
        #'.azure_stack_ou': f'OU={ad.azure_stack_ou},DC=' + ad.domain.replace('.', ',DC='),
        #'.domain': ad.domain,
        #'.azure_stack_lcm_user': ad.azure_stack_lcm_user.split('@')[0],
        '.hostname': v.name,
        '.interface_1_identifier': f'SlotID {vnics[0]} Port 1',
        '.interface_2_identifier': f'SlotID {vnics[0]} Port 2',
        '.ntp_server': kwargs.ntp_servers[0],
        '.organization': kwargs.imm_dict.wizard.azure_stack[0].organization,
        #'.secure.azure_stack_lcm_password': kwargs.azure_stack_lcm_password,
        '.secure.local_administrator_password': kwargs.local_administrator_password,
        #'.macAddressNic1_dash_format': mac_a,
        #'.macAddressNic2_dash_format': mac_b
        # Language Pack/Localization
        '.input_locale': kwargs.language.input_locale,
        '.language_pack': kwargs.language.ui_language,
        '.layered_driver': kwargs.language.layered_driver,
        '.secondary_language': kwargs.language.secondary_language}
        # Timezone Configuration
        #'.disable_daylight_savings': kwargs.disable_daylight,
        #'.timezone': kwargs.windows_timezone}
    answers_dict = dict(sorted(answers_dict.items()))
    for e in ['layered_driver', 'secondary_language']:
        if kwargs.language[e] == '': answers_dict.pop(f'.{e}')
    for key,value in answers_dict.items(): api_body['AdditionalParameters'].append(installation_body_os_placeholders(key, value))
    return api_body

#=============================================================================
# Function - OS Install Custom Template Parameters Map
#=============================================================================
def installation_body_os_placeholders(key, value):
    if 'secure' in key: secure = True
    else: secure = False
    if len(value) == 0: is_set = False
    else: is_set = True
    parameters = {
        'ClassId': 'os.PlaceHolder',
        'IsValueSet': is_set,
        'ObjectType': 'os.PlaceHolder',
        'Type': {
            'ClassId': 'workflow.PrimitiveDataType',
            'Default': {
                'ClassId': 'workflow.DefaultValue',
                'IsValueSet': False,
                'ObjectType': 'workflow.DefaultValue',
                'Override': False,
                'Value': None},
            'Description': '',
            'DisplayMeta': {
                'ClassId': 'workflow.DisplayMeta',
                'InventorySelector': True,
                'ObjectType': 'workflow.DisplayMeta',
                'WidgetType': 'None'},
            'InputParameters': None,
            'Label': key,
            'Name': key,
            'ObjectType': 'workflow.PrimitiveDataType',
            'Properties': {
                'ClassId': 'workflow.PrimitiveDataProperty',
                'Constraints': {
                    'ClassId': 'workflow.Constraints',
                    'EnumList': [],
                    'Max': 0,
                    'Min': 0,
                    'ObjectType': 'workflow.Constraints',
                    'Regex': ''},
                'InventorySelector': [],
                'ObjectType': 'workflow.PrimitiveDataProperty',
                'Secure': secure,
                'Type': 'string'},
            'Required': False},
        'Value': value}
    return parameters

#=============================================================================
# Function - Build api_body for Operating System Installation - VMware
#=============================================================================
def installation_body_vmware(v, kwargs):
    api_body = {
        'AdditionalParameters': None,
        'Answers': {
            'AlternateNameServers': [],
            'Hostname': f'{v.name}.{kwargs.dns_domains[0]}',
            'IpConfigType': 'static',
            'IpConfiguration': {
                'IpV4Config': {
                    'Gateway': v.inband.gateway,
                    'IpAddress': v.inband.ip,
                    'Netmask': v.inband.netmask,
                    'ObjectType': 'comm.IpV4Interface'},
                'ObjectType': 'os.Ipv4Configuration'},
            'IsRootPasswordCrypted': False,
            'Nameserver': kwargs.dns_servers[0],
            'NetworkDevice': kwargs.mac_a,
            'ObjectType': 'os.Answers',
            'RootPassword': kwargs.vmware_esxi_password,
            'Source': 'Template'},
        'ConfigurationFile': {'Moid': kwargs.os_cfg_moid, 'ObjectType': 'os.ConfigurationFile'},
        'Image': {'Moid': kwargs.os_sw_moid, 'ObjectType': 'softwarerepository.OperatingSystemFile'},
        'InstallMethod': 'vMedia',
        'InstallTarget': {},
        'ObjectType': 'os.Install',
        'Organization': {'Moid': kwargs.org_moid, 'ObjectType': 'organization.Organization'},
        'OsduImage': {'Moid': kwargs.scu_moid, 'ObjectType': 'firmware.ServerConfigurationUtilityDistributable'},
        'OverrideSecureBoot': True,
        'Server': {'Moid': v.hardware_moid, 'ObjectType': v.object_type}}
    if len(kwargs.dns_servers) > 0:
        api_body['Answers']['AlternateNameServers'] = [kwargs.dns_servers[1]]
    return api_body

#=============================================================================
# Function - Prompt User for the Intersight Configurtion
#=============================================================================
def intersight_config(kwargs):
    kwargs.jdata = DotMap()
    if kwargs.args.intersight_api_key_id == None:
        kwargs.sensitive_var = 'intersight_api_key_id'
        kwargs = sensitive_var_value(kwargs)
        kwargs.args.intersight_api_key_id = kwargs.var_value
    #=========================================================================
    # Prompt User for Intersight SecretKey File
    #=========================================================================
    secret_path = kwargs.args.intersight_secret_key
    if not re.search('BEGIN RSA PRIVATE KEY.*END RSA PRIVATE KEY', secret_path):
        secret_loop = False
        while secret_loop == False:
            valid = False
            if secret_path == None:
                varName = 'intersight_secret_key'
                pcolor.Cyan(f'\n{"-"*108}\n\n  The Script did not find {varName} as an `environment` variable.')
                pcolor.Cyan(f'  To not be prompted for the value of {varName} each time\n  add the following to your local environemnt:')
                pcolor.Cyan(f'    - Linux: export {varName}="{varName}_value"')
                pcolor.Cyan(f'    - Windows: $env:{varName}="{varName}_value"')
                secret_path = ''
            if '~' in secret_path: secret_path = os.path.expanduser(secret_path)
            if not secret_path == '':
                if not os.path.isfile(secret_path): pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!! intersight_secret_key not found.')
                else:
                    def key_error(secret_path):
                        pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  Path: {secret_path}\n   does not seem to contain a Valid PEM Secret Key.')
                        pcolor.Red(f'\n{"-"*108}\n')
                        len(False); sys.exit(1)
                    if 'EC PRIVATE KEY' in open(secret_path).read():
                        try: SigningKey.from_pem(open(secret_path).read())
                        except Exception as e: pcolor.Red(e); key_error(secret_path)
                    elif 'RSA PRIVATE KEY' in open(secret_path).read():
                        try: RSA.RsaKey.has_private(RSA.import_key(open(secret_path).read()))
                        except Exception as e: pcolor.Red(e); key_error(secret_path)
                    else: key_error(secret_path)
                    kwargs.args.intersight_secret_key = secret_path; secret_loop = True; valid = True
            if not valid == True:
                kwargs.jdata = DotMap(
                    type = 'string', minLength = 2, maxLength = 1024, pattern = '.*', title = 'Intersight',
                    description = 'Intersight Secret Key File Location.',
                    default     = os.path.join(kwargs.home, 'Downloads', 'SecretKey.txt'))
                secret_path = variable_prompt(kwargs)
    #=========================================================================
    # Prompt User for Intersight FQDN
    #=========================================================================
    valid = False
    while valid == False:
        varValue = kwargs.args.intersight_fqdn
        if not varValue == None:
            varName = 'Intersight FQDN'
            if re.search(r'^[a-zA-Z0-9]{1,4}:', varValue): valid = validating.ip_address(varName, varValue)
            elif re.search(r'[a-zA-Z]', varValue): valid = validating.dns_name(varName, varValue)
            elif re.search(r'^([0-9]{1,3}\.){3}[0-9]{1,3}$', varValue): valid = validating.ip_address(varName, varValue)
            else: pcolor.Red(f'\n{"-"*108}\n\n  "{varValue}" is not a valid address.\n\n{"-"*108}\n')
        if valid == False:
            kwargs.jdata = kwargs.ezdata.ntp.allOf[1].properties.ntp_servers['items']
            kwargs.jdata.update(DotMap(description = 'Hostname of the Intersight FQDN',
                                       default = 'intersight.com', title = 'Intersight FQDN'))
            kwargs.args.intersight_fqdn = variable_prompt(kwargs)
            valid = True
    # Return kwargs
    return kwargs

#=============================================================================
# Function - Load Previous YAML Files
#=============================================================================
def load_previous_configurations(kwargs):
    ezvars    = kwargs.ezdata.ezimm_class.properties
    vclasses  = kwargs.ezdata.ezimm_class.properties.classes.enum
    dir_check = 0
    if os.path.isdir(kwargs.args.dir):
        dir_list = os.listdir(kwargs.args.dir)
        for i in dir_list:
            if i == 'templates':  dir_check += 1
            elif i == 'policies': dir_check += 1
            elif i == 'pools':    dir_check += 1
            elif i == 'profiles': dir_check += 1
            elif i == 'wizard': dir_check += 1
    if dir_check > 0:
        for item in vclasses:
            dest_dir = ezvars[item].directory
            if os.path.isdir(os.path.join(kwargs.args.dir, dest_dir)):
                dir_list = os.listdir(os.path.join(kwargs.args.dir, dest_dir))
                for i in dir_list:
                    if os.path.isfile(os.path.join(kwargs.args.dir, dest_dir, i)) and re.search('.*yaml$', i):
                        yfile = open(os.path.join(kwargs.args.dir, dest_dir, i), 'r')
                        data = yaml.safe_load(yfile)
                        if not data == None:
                            for key, value in data.items():
                                if not kwargs.imm_dict.orgs.get(key): kwargs.imm_dict.orgs[key] = {}
                                if type(value) == dict:
                                    for k, v in value.items():
                                        if not kwargs.imm_dict.orgs[key].get(k): kwargs.imm_dict.orgs[key][k] = {}
                                        kwargs.imm_dict.orgs[key][k].update(deepcopy(v))
                                elif type(value) == str or type(value) == bool or type(value) == list: kwargs.imm_dict.orgs[key] = value
                                else:
                                    pcolor.Yellow(f'failed to match type {type(value)}')
                                    len(False); sys.exit(1)
    # Return kwargs
    return kwargs

#=============================================================================
# Function - Local User Policy - Users
#=============================================================================
def local_users_function(kwargs):
    loop_count = 1
    kwargs.local_users = []
    valid_config = False
    while valid_config == False:
        #=========================================================================
        # Loop Through Local User Atttributes
        #=========================================================================
        attributes    = DotMap()
        attribute_list= list(kwargs.ezdata['local_user.users'].properties.keys())
        attribute_list.remove('password')
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['local_user.users'].properties[e]
            kwargs.jdata.multi_select = False
            attributes[e] = variable_prompt(kwargs)
        if kwargs.enforce_strong_password == True:
            pcolor.LightPurple(kwargs.ezdata.local_user.password_properties.enforce_strong_password.description)
        kwargs.sensitive_var = f'local_user_password_{loop_count}'
        kwargs = sensitive_var_value(kwargs)
        attributes.password = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #=========================================================================
        # Show User Configuration
        #=========================================================================
        pcolor.Green(f'\n{"-"*108}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=yaml_dumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*108}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'Local User'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.local_users.append(attributes)
        loop_count += 1
        valid_config = kwargs.configure_additional
    return kwargs

#=============================================================================
# Function - Merge Easy IMM Repository to Dest Folder
#=============================================================================
def merge_easy_imm_repository(kwargs):
    # Download the Easy IMM Comprehensive Example Base Repo
    baseRepo= kwargs.args.dir
    tfe_dir = 'tfe_modules'
    git_url = 'https://github.com/terraform-cisco-modules/easy-imm-comprehensive-example'
    if not os.path.isdir(tfe_dir): os.mkdir(tfe_dir); Repo.clone_from(git_url, tfe_dir)
    if not os.path.isfile(os.path.join(tfe_dir, 'README.md')): Repo.clone_from(git_url, tfe_dir)
    else: g = cmd.Git(tfe_dir); g.pull()
    if not os.path.isdir(baseRepo): os.mkdir(baseRepo)
    # Now Loop over the folders and merge the module files
    for folder in ['defaults', '']:
        if folder == 'defaults':
            dest_dir = os.path.join(baseRepo, folder); src_dir = os.path.join(tfe_dir, 'defaults')
            if not os.path.isdir(dest_dir): os.mkdir(dest_dir)
        else: dest_dir = os.path.join(baseRepo); src_dir = os.path.join(tfe_dir)
        copy_files = os.listdir(src_dir)
        for fname in copy_files:
            if not os.path.isdir(os.path.join(src_dir, fname)): shutil.copy2(os.path.join(src_dir, fname), dest_dir)

#=============================================================================
# Function - Message for Invalid List Selection
#=============================================================================
def message_invalid_selection():
    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!! Invalid Selection.  Please Select a valid Option from the List.')
    pcolor.Red(f'\n{"-"*108}\n')

#=============================================================================
# Function - Message for Invalid Selection Y or N
#=============================================================================
def message_invalid_y_or_n(length):
    if length == 'short': dash_rep = '-'*54
    else: dash_rep = '-'*108
    pcolor.Red(f'\n{dash_rep}\n\n  !!!Error!!! Invalid Value.  Please enter `Y` or `N`.')
    pcolor.Red(f'\n{dash_rep}\n')

#=============================================================================
# Function - Message Invalid FCoE VLAN
#=============================================================================
def message_fcoe_vlan(fcoe_id, vlan_policy):
    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  The FCoE VLAN `{fcoe_id}` is already assigned to the VLAN Policy')
    pcolor.Red(f'  {vlan_policy}.  Please choose a VLAN id that is not already in use.')
    pcolor.Red(f'\n{"-"*108}\n')

#=============================================================================
# Function(s) - Message Invalid Native VLAN
#=============================================================================
def message_invalid_native_vlan(nativeVlan, VlanList):
    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  The Native VLAN `{nativeVlan}` was not in the VLAN Policy List.')
    pcolor.Red(f'  VLAN Policy List is: "{VlanList}"')
    pcolor.Red(f'\n{"-"*108}\n')

#=============================================================================
# Function - Message Invalid VLAN/VSAN
#=============================================================================
def message_invalid_vxan():
    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  Invalid Entry.  Please Enter a valid ID in the range of 1-4094.')
    pcolor.Red(f'\n{"-"*108}\n')

#=============================================================================
# Function - Message Invalid VLAN
#=============================================================================
def message_invalid_vsan_id(vsan_policy, vsan_id, vsan_list):
    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  The VSAN `{vsan_id}` is not in the VSAN Policy `{vsan_policy}`.')
    pcolor.Red(f'  Options are: {vsan_list}.\n\n{"-"*108}\n')

#=============================================================================
# Function - Message Starting Over
#=============================================================================
def message_starting_over(policy_type):
    pcolor.Yellow(f'\n{"-"*54}\n\n  Starting `{policy_type}` Section over.')
    pcolor.Yellow(f'\n{"-"*54}\n')

#=============================================================================
# Function - Change Policy Description to Sentence
#=============================================================================
def mod_pol_description(pol_description):
    pdescr = str.title(pol_description.replace('_', ' '))
    pdescr = (((pdescr.replace('Fiattached', 'FIAttached')).replace('Imc', 'IMC')).replace('Mac', 'MAC')).replace('Uuid', 'UUID')
    pdescr = (((pdescr.replace('Iscsi', 'iSCSI')).replace('Fc', 'FC')).replace('San', 'SAN')).replace('Lan', 'LAN')
    pdescr = (((pdescr.replace('Ipmi', 'IPMI')).replace('Ip', 'IP')).replace('Iqn', 'IQN')).replace('Ldap', 'LDAP')
    pdescr = (((pdescr.replace('Ntp', 'NTP')).replace('Sd', 'SD')).replace('Smtp', 'SMTP')).replace('Snmp', 'SNMP')
    pdescr = (((pdescr.replace('Ssh', 'SSH')).replace('Wwnn', 'WWNN')).replace('Wwpn', 'WWPN')).replace('Vsan', 'VSAN')
    pdescr = (((pdescr.replace('Vnics', 'vNICs')).replace('Vhbas', 'vHBAs')).replace('Vlan', 'VLAN')).replace('Os Install', 'OS')
    return pdescr

#=============================================================================
# Function - Change Policy Description to Sentence
#=============================================================================
def name_prefix_suffix(policy, kwargs):
    if   re.search('^ip|iqn|mac|resource|uuid|wwnn|wwpn$', policy): ptype = 'pools'
    elif re.search('^profiles.(chassis|domain|server)$', policy): ptype = 'profiles'
    elif re.search('^templates.(chassis|domain|server)$', policy): ptype = 'templates'
    else: ptype = 'policies'
    args  = DotMap(name_prefix = '', name_suffix = '')
    pkeys = list(kwargs.imm_dict.orgs[kwargs.org][ptype].keys())
    for e in ['name_prefix', 'name_suffix']:
        if e in pkeys:
            nkeys = list(kwargs.imm_dict.orgs[kwargs.org][ptype][e].keys())
            if policy in nkeys:
                if len(kwargs.imm_dict.orgs[kwargs.org][ptype][e][policy]) > 0:
                    args[e] = kwargs.imm_dict.orgs[kwargs.org][ptype][e][policy]
            if args[e] == '':
                if 'default' in nkeys:
                    if len(kwargs.imm_dict.orgs[kwargs.org][ptype][e]['default']) > 0:
                        args[e] = kwargs.imm_dict.orgs[kwargs.org][ptype][e]['default']
    return args.name_prefix, args.name_suffix

#=============================================================================
# Function - Naming Rule
#=============================================================================
def naming_rule(name_prefix, name_suffix, org):
    if not name_prefix == '':  name = f'{name_prefix}_{name_suffix}'
    else: name = f'{org}_{name_suffix}'
    return name

#=============================================================================
# Function - Naming Rule Fabric Policy
#=============================================================================
def naming_rule_fabric(loop_count, name_prefix, org):
    letter = chr(ord('@')+loop_count+1)
    if not name_prefix == '':   name = f'{name_prefix}-{letter.lower()}'
    elif not org == 'default':  name = f'{org}-{letter.lower()}'
    else: name = f'fabric-{letter.lower()}'
    return name

#=============================================================================
# Function - Build api_body for OS Configuration Item
#=============================================================================
def os_configuration_file(kwargs):
    api_body = {
        'Catalog': kwargs.org_catalog_moid,
        'Description': '',
        'Distributions': [{'Moid': kwargs.distribution_moid, 'ObjectType': 'hcl.OperatingSystem'}],
        'FileContent': kwargs.file_content,
        'Internal': False,
        'Name': kwargs.os_config_template,
        'ObjectType': 'os.ConfigurationFile',
        'Tags': [e.toDict() for e in kwargs.ez_tags]}
    return api_body

#=============================================================================
# Function - Determine Adapter PCI Slot
#=============================================================================
def pci_slot(element):
    if element.PciSlot == 'SlotID:0-MLOM' or 'MLOM' in element.Model:  pci_slot = 'MLOM'
    elif not 'MEZZ' in element.PciSlot and 'SlotID' in element.PciSlot:
        pci_slot = re.search('SlotID:(\\d)', element.PciSlot).group(1)
    elif re.search('\\d', str(element.PciSlot)): pci_slot = int(element.PciSlot)
    elif re.search('L', str(element.PciSlot)): pci_slot = 'LOM'
    else: pci_slot = e.AdapterId
    return pci_slot

#=============================================================================
# Function - Get Policies from Dictionary
#=============================================================================
def policies_parse(ptype, policy_type, kwargs):
    org  = kwargs.org
    kwargs.policies = []
    if not kwargs.imm_dict.orgs[org].get(ptype) == None:
        if not kwargs.imm_dict.orgs[org][ptype].get(policy_type) == None:
            kwargs.policies = {policy_type:kwargs.imm_dict.orgs[org][ptype][policy_type]}
        else: kwargs.policies = {policy_type:{}}
    else: kwargs.policies = {policy_type:{}}
    return kwargs

#=============================================================================
# Function - Validate input for each method
#=============================================================================
def process_kwargs(kwargs):
    #=========================================================================
    # Validate User Input
    #=========================================================================
    json_data = kwargs['validateData']
    validate_args(kwargs)
    error_count = 0; error_list = []
    optional_args = json_data['optional_args']
    required_args = json_data['required_args']
    for item in required_args:
        if item not in kwargs['var_dict'].keys(): error_count =+ 1; error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following REQUIRED Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise insufficient_args(error_)
    #=========================================================================
    # Load all optional args values from kwargs
    #=========================================================================
    error_count = 0; error_list = []
    for item in optional_args:
        if item not in kwargs['var_dict'].keys(): error_count =+ 1; error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following Optional Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise insufficient_args(error_)
    #=========================================================================
    # Load all required args values from kwargs
    #=========================================================================
    error_count = 0; error_list = []
    for item in kwargs['var_dict']:
        if item in required_args.keys():
            required_args[item] = kwargs['var_dict'][item]
            if required_args[item] == None: error_count =+ 1; error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "%s"\n\n****End ERROR****\n' % (error_list)
        raise insufficient_args(error_)
    for item in kwargs['var_dict']:
        if item in optional_args.keys(): optional_args[item] = kwargs['var_dict'][item]
    # Combine option and required dicts for Jinja template render
    pol_vars = {**required_args, **optional_args}
    return(pol_vars)

#=============================================================================
# Function - Read Excel Workbook Data
#=============================================================================
def read_in(excel_workbook, kwargs):
    try:
        kwargs['wb'] = load_workbook(excel_workbook)
        pcolor.Cyan('Workbook Loaded.')
    except Exception as e:
        pcolor.Red(f'\n{"-"*108}\n\n  Something went wrong while opening the workbook - {excel_workbook}... ABORT!')
        pcolor.Red(f'\n{"-"*108}\n')
        sys.exit(e)
    return kwargs

#=============================================================================
# Remove Duplicate Entries in Python Dictionary
#=============================================================================
def remove_duplicates(orgs, plist, kwargs):
    idict = {}
    for org in orgs:
        idict[org] = {}
        for d in plist:
            idict[org][d] = {}
            for e in list(kwargs.imm_dict.orgs[org][d].keys()):
                if type(kwargs.imm_dict.orgs[org][d][e]) == list:
                    idict[org][d][e] = []
                    for i in kwargs.imm_dict.orgs[org][d][e]:
                        i = dict(sorted(i.items()))
                        if not i in idict[org][d][e]: idict[org][d][e].append(i)
                    kwargs.imm_dict.orgs[org][d][e] = deepcopy(idict[org][d][e])
            kwargs.imm_dict.orgs[org][d] = DotMap(sorted(kwargs.imm_dict.orgs[org][d].items()))
    return kwargs

#=============================================================================
# Function - Prompt User for Sensitive Values
#=============================================================================
def sensitive_var_value(kwargs):
    sensitive_var = kwargs.sensitive_var
    undefined = False
    if re.search('^undefined_', sensitive_var):
        undefined = True
        sensitive_var = sensitive_var.replace('undefined_', '')
    #=======================================================================================================
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #=======================================================================================================
    if os.environ.get(sensitive_var) is None:
        pcolor.Cyan(f'\n{"-"*108}\n')
        pcolor.Cyan(f'  The Script did not find {sensitive_var} as an `environment` variable.')
        pcolor.Cyan(f'  To not be prompted for the value of `{kwargs.sensitive_var}` each time')
        pcolor.Cyan(f'  add the following to your local environemnt:\n')
        pcolor.Cyan(f'    - Linux: export {sensitive_var}="{kwargs.sensitive_var}_value"')
        pcolor.Cyan(f'    - Windows: $env:{sensitive_var}="{kwargs.sensitive_var}_value"\n')
    if os.environ.get(sensitive_var) is None and kwargs.sensitive_var == 'ipmi_key':
        pcolor.Cyan(f'\n{"-"*108}\n\n  The ipmi_key Must be in Hexidecimal Format [a-fA-F0-9]')
        pcolor.Cyan(f'  and no longer than 40 characters.\n')
    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '': valid = True
        valid = False
        while valid == False:
            if kwargs.get('multi_line_input'):
                pcolor.LightGray(f'Enter the value for {kwargs.sensitive_var}:')
                lines = []
                while True:
                    line = stdiomask.getpass(prompt='')
                    if line: lines.append(line)
                    else: break
                if not re.search('(certificate|private_key)', sensitive_var): secure_value = '\\n'.join(lines)
                else: secure_value = '\n'.join(lines)
            else:
                valid_pass = False
                while valid_pass == False:
                    password1 = stdiomask.getpass(prompt=f'Enter the value for {kwargs.sensitive_var}: ')
                    password2 = stdiomask.getpass(prompt=f'Re-Enter the value for {kwargs.sensitive_var}: ')
                    if password1 == password2: secure_value = password1; valid_pass = True
                    else: pcolor.Red('!!! ERROR !!! Sensitive Values did not match.  Please re-enter...')
            #=========================================================================
            # Validate Sensitive Passwords
            #=========================================================================
            #cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            if re.search('(certificate|private_key)', sensitive_var):
                try:
                    if re.search('certficate', sensitive_var):
                        pem = crypto.load_certificate(crypto.FILETYPE_PEM, open(secure_value).read())
                        expiration_date = pem.get_notAfter().decode('utf-8')
                        formatted_date  = datetime.strptime(expiration_date, '%Y%m%d%H%M%SZ')
                        pcolor.Cyan(f'{sensitive_var} Certficate Expiration is {formatted_date}')
                        pem = True
                    else:
                        pem = RSA.RsaKey.has_private(RSA.import_key(open(secure_value).read()))
                except Exception as e:
                    pcolor.Red(e)
                    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  Path: {sensitive_var}\n   does not seem to be valid.')
                    pcolor.Red(f'\n{"-"*108}\n')
                    len(False); sys.exit(1)
                if pem == True: valid = True
                #if re.search(cert_regex, secure_value): valid = True
                else:
                    pcolor.Red(f'\n{"-"*108}\n\n  !!!Error!!!\n  Path: {sensitive_var}\n   does not seem to be valid.')
                    pcolor.Red(f'\n{"-"*108}\n')
            elif undefined == True: valid = True
            elif re.search('intersight_api_key_id', sensitive_var):
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.intersight_api_key_id
                valid = validate_sensitive(secure_value, kwargs)
            elif 'bind' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.ldap_binding_password
                valid = validate_sensitive(secure_value, kwargs)
            elif 'community' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.snmp_community_string
                valid = validate_sensitive(secure_value, kwargs)
            elif 'ipmi_key' in sensitive_var: valid = validating.ipmi_key_check(secure_value)
            elif 'iscsi_boot' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.iscsi_boot_password
                valid = validate_sensitive(secure_value, kwargs)
            elif re.search('(local_user|root|ucs(_central)?)_password', sensitive_var):
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.local_user_password
                if kwargs.enforce_strong_password == True:
                    kwargs.jdata.maxLength = 20
                    valid = validate_strong_password(secure_value, kwargs)
                else: valid = validate_sensitive(secure_value, kwargs)
            elif 'persistent_passphrase' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.persistent_passphrase
                valid = validate_sensitive(secure_value, kwargs)
            elif 'snmp' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.snmp_password
                valid = validate_sensitive(secure_value, kwargs)
            elif 'vmedia' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.properties.vmedia_password
                valid = validate_sensitive(secure_value, kwargs)
        #=========================================================================
        # Add Policy Variables to imm_dict
        #=========================================================================
        if kwargs.get('org'):
            org = kwargs.org
            if not kwargs.imm_dict.orgs.get(org):
                kwargs.imm_dict.orgs[org] = DotMap()
                if not kwargs.imm_dict.orgs[org].get('sensitive_vars'): kwargs.imm_dict.orgs[org].sensitive_vars = []
                kwargs.imm_dict.orgs[org].sensitive_vars.append(sensitive_var)
        #=========================================================================
        # Add the Variable to the Environment
        #=========================================================================
        os.environ[sensitive_var] = '%s' % (secure_value)
        kwargs.var_value = secure_value
    else:
        #=========================================================================
        # Add the Variable to the Environment
        #=========================================================================
        if not kwargs.get('multi_line_input'): kwargs.var_value = os.environ.get(sensitive_var)
        else: kwargs.var_value = (os.environ.get(sensitive_var)).replace('\n', '\\n')
    return kwargs

#=============================================================================
# Function - Wizard for SNMP Trap Servers
#=============================================================================
def snmp_trap_servers(kwargs):
    loop_count = 1
    kwargs.snmp_traps = []
    valid_config = False
    while valid_config == False:
        #=========================================================================
        # Loop Through SNMP Trap Server Atttributes
        #=========================================================================
        attributes= DotMap()
        attribute_list = list(kwargs.ezdata['snmp.snmp_trap_destinations'].properties.keys())
        if len(kwargs.snmp_users) > 0:
            for e in ['community_string', 'trap_type', 'user']: attribute_list.remove(e)
        else: attribute_list.pop('user')
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties[e]
            kwargs.jdata.multi_select = False
            attributes[e] = variable_prompt(kwargs)
        if len(kwargs.snmp_users) > 0:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties['user']
            kwargs.jdata.enum = [e.name for e in kwargs.snmp_users]
            kwargs.jdata.multi_select = False
            attributes.user = variable_prompt(kwargs)
        else:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties['trap_type']
            kwargs.jdata.multi_select = False
            attributes.trap_type = variable_prompt(kwargs)
            kwargs.sensitive_var = f'snmp_trap_community_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.community_string = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #=========================================================================
        # Show User Configuration
        #=========================================================================
        pcolor.Green(f'\n{"-"*108}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=yaml_dumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*108}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'SNMP Trap Destination'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.snmp_traps.append(attributes)
        loop_count += 1
        valid_config = kwargs.configure_additional
    return kwargs

#=============================================================================
# Function - Wizard for SNMP Users
#=============================================================================
def snmp_users(kwargs):
    loop_count = 1
    kwargs.snmp_users = []
    valid_users = False
    while valid_users == False:
        #=========================================================================
        # Loop Through SNMP User Atttributes
        #=========================================================================
        attributes    = DotMap()
        attribute_list= list(kwargs.ezdata['snmp.snmp_users'].properties.keys())
        for e in ['auth_password', 'privacy_password', 'privacy_type']: attribute_list.remove(e)
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_users'].properties[e]
            attributes[e] = variable_prompt(kwargs)
        if attributes.security_level == 'AuthPriv':
            kwargs.jdata = kwargs.ezdata['snmp.snmp_users'].properties.privacy_type
            attributes.privacy_type = variable_prompt(kwargs)
        if re.search('Auth(No)?Priv', attributes.security_level):
            kwargs.sensitive_var = f'snmp_auth_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.auth_password = loop_count
        if re.search('AuthPriv', attributes.security_level):
            kwargs.sensitive_var = f'snmp_privacy_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.privacy_password = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #=========================================================================
        # Show User Configuration
        #=========================================================================
        pcolor.Green(f'\n{"-"*108}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=yaml_dumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*108}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'SNMP User'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.snmp_users.append(attributes)
        loop_count += 1
        valid_users = kwargs.configure_additional
    return kwargs

#=============================================================================
# Function - Define stdout_log output
#=============================================================================
def stdout_log(ws, row_num):
    if log_level == 0: return
    elif ((log_level == (1) or log_level == (2)) and (ws) and (row_num is None)) and row_num == 'begin':
        pcolor.Cyan(f'\n{"-"*108}\n\n   Begin Worksheet "{ws.title}" evaluation...')
        pcolor.Cyan(f'\n{"-"*108}\n')
    elif (log_level == (1) or log_level == (2)) and row_num == 'end':
        pcolor.Cyan(f'\n{"-"*108}\n\n   Completed Worksheet "{ws.title}" evaluation...')
        pcolor.Cyan(f'\n{"-"*108}\n')
    elif log_level == (2) and (ws) and (row_num is not None):
        pcolor.Cyan(f'    - Evaluating Row{" "*(4-len(row_num))}{row_num}...')
    else: return

#=============================================================================
# Function - Wizard for Syslog Servers
#=============================================================================
def syslog_servers(kwargs):
    kwargs.remote_logging = []
    loop_count = 0
    valid_config = False
    while valid_config == False:
        if loop_count < 2:
            #=========================================================================
            # Loop Through SNMP Trap Server Atttributes
            #=========================================================================
            attributes= DotMap()
            attribute_list = list(kwargs.ezdata['syslog.remote_logging'].properties.keys())
            for e in attribute_list:
                kwargs.jdata = kwargs.ezdata['syslog.remote_logging'].properties[e]
                kwargs.jdata.multi_select = False
                attributes[e] = variable_prompt(kwargs)
            #=========================================================================
            # Show User Configuration
            #=========================================================================
            pcolor.Green(f'\n{"-"*108}\n')
            pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=yaml_dumper, default_flow_style=False), " "*3, predicate=None))
            pcolor.Green(f'\n{"-"*108}\n')
            #======================================================================
            # * Prompt User to Accept Configuration, If Accepted add to Dictionary
            # * If User Selects to, Configure Additional
            #======================================================================
            if loop_count == 0: kwargs.yes_or_no  = 'Y'
            else: kwargs.yes_or_no  = 'N'
            kwargs.policy_type= 'Syslog Remote Servers'
            kwargs = exit_confirm_loop(kwargs)
            if kwargs.accept_configuration == True: kwargs.remote_logging.append(attributes)
            loop_count += 1
            valid_config = kwargs.configure_additional
        else: valid_config = True
    return kwargs

#=============================================================================
# Function - Create a List of Subnet Hosts
#=============================================================================
def subnet_list(kwargs):
    if kwargs.ip_version == 'v4': prefix = kwargs.subnetMask
    else: prefix = kwargs.prefix
    gateway = kwargs.defaultGateway
    return list(ipaddress.ip_network(f'{gateway}/{prefix}', strict=False).hosts())

#=============================================================================
# Function - Format Terraform Files
#=============================================================================
def terraform_fmt(folder):
    # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
    pcolor.Cyan(f'\n{"-"*108}\n')
    pcolor.Cyan(f'  Running "terraform fmt" in folder "{folder}",')
    pcolor.Cyan(f'  to correct variable formatting!')
    pcolor.Cyan(f'\n{"-"*108}\n')
    p = subprocess.Popen(['terraform', 'fmt', folder], stdout = subprocess.PIPE, stderr = subprocess.PIPE)
    pcolor.Cyan('Format updated for the following Files:')
    for line in iter(p.stdout.readline, b''):
        line = line.decode('utf-8')
        line = line.strip()
        pcolor.Cyan(f'- {line}')

#=============================================================================
# Function to Pull Latest Versions of Providers
#=============================================================================
def terraform_provider_config(kwargs):
    url_list = [
        'https://github.com/CiscoDevNet/terraform-provider-intersight/tags/',
        'https://github.com/hashicorp/terraform/tags',
        'https://github.com/netascode/terraform-provider-utils/tags/']
    for url in url_list:
        # Get the Latest Release Tag for the Provider
        r = requests.get(url, stream=True)
        repoVer = 'BLANK'
        stringMatch = False
        while stringMatch == False:
            for line in r.iter_lines():
                toString = line.decode('utf-8')
                if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                    repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                    break
            stringMatch = True
        # Make sure the latest_versions Key exists
        if kwargs.get('latest_versions') == None: kwargs.latest_versions = {}
        # Set Provider Version
        if   'intersight' in url: kwargs.latest_versions.intersight_provider_version = repoVer
        elif 'netascode' in url:  kwargs.latest_versions.utils_provider_version = repoVer
        else: kwargs.latest_versions.terraform_version = repoVer
    # Return kwargs
    return kwargs
    
#=============================================================================
# Function - Test Repo URL for File
#=============================================================================
def test_repository_url(repo_url):
    try: requests.head(repo_url, allow_redirects=True, verify=False, timeout=10)
    except requests.RequestException as e:
        pcolor.Red(f'\n{"-"*108}\n')
        pcolor.Red(f'!!! ERROR !!!\n  Exception when calling {repo_url}:\n {e}\n')
        pcolor.Red(f'Please Validate the Software Repository is setup properly.  Exiting...')
        pcolor.Red(f'\n{"-"*108}\n')
        len(False); sys.exit(1)

#=============================================================================
# Function - Prompt User for Sensitive Variables
#=============================================================================
def tfc_sensitive_variables(var_value, kwargs):
    pol_vars = DotMap(Variable = var_value)
    pol_vars.Description = (''.join(var_value.split('_'))).title()
    pol_vars.Description = mod_pol_description(pol_vars.Description)
    kwargs = sensitive_var_value(kwargs)
    pol_vars.varValue = kwargs.var_value
    pol_vars.varId = var_value
    pol_vars.varKey = var_value
    pol_vars.Sensitive = True
    pcolor.Cyan(f'  * Adding "{pol_vars.description}" to "{kwargs.workspaceName}"')
    return pol_vars

#=============================================================================
# Function - Prompt User for Chassis/Server Serial Numbers
#=============================================================================
def ucs_serial(kwargs):
    baseRepo    = kwargs.args.dir
    device_type = kwargs.device_type
    org         = kwargs.org
    yaml_file   = kwargs.yaml_file
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n{"-"*108}\n')
        pcolor.Cyan(f'  Note: If you do not have the Serial Number(s) at this time you can manually add it to:')
        pcolor.Cyan(f'    - {os.path.join(baseRepo, org, "profiles", f"{yaml_file}.yaml")}')
        pcolor.Cyan(f'      file later.')
        pcolor.Cyan(f'\n{"-"*108}\n')
        serial = input(f'What is the Serial Number of the {device_type}? [press enter to skip]: ')
        if serial == '': serial = 'unknown'; valid = True
        elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', serial): valid = True
        else:
            pcolor.Red(f'\n{"-"*108}\n  Error!! Invalid Serial Number.  "{serial}" is not a valid serial.')
            pcolor.Red(f'\n{"-"*108}\n')
    return serial

#=============================================================================
# Function - Prompt User for Domain Serial Numbers
#=============================================================================
def ucs_domain_serials(kwargs):
    baseRepo = kwargs['args'].dir
    org = kwargs['org']
    pcolor.Cyan(f'\n{"-"*108}\n')
    pcolor.Cyan(f'  Note: If you do not have the Serial Numbers at this time you can manually add them here:\n')
    pcolor.Cyan(f'    * {os.path.join(baseRepo, "profiles", f"domain.yaml")}\n')
    pcolor.Cyan(f'  After the Wizard has completed.')
    pcolor.Cyan(f'\n{"-"*108}\n')
    valid = False
    while valid == False:
        pol_vars = {}
        fabrics = ['A','B']
        for x in fabrics:
            pol_vars[f'serial_{x}'] = input(f'What is the Serial Number of Fabric {x}? [press enter to skip]: ')
            if pol_vars[f'serial_{x}'] == '':
                pol_vars[f'serial_{x}'] = 'unknown'
                valid = True
            elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', pol_vars[f'serial_{x}']):
                valid = True
            else:
                pcolor.Red(f'\n{"-"*108}\n')
                pcolor.Red('  Error!! Invalid Serial Number.  "{}" is not a valid serial.').format(pol_vars[f'serial_{x}'])
                pcolor.Red(f'\n{"-"*108}\n')
    serials = [pol_vars['serial_A'], pol_vars['serial_B']]
    return serials

#=============================================================================
# Function to Validate Worksheet User Input
#=============================================================================
def validate_args(json_data, kwargs):
    json_data = kwargs['validateData']
    for i in json_data['required_args']:
        if json_data[i]['type'] == 'boolean':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.boolean(i, kwargs)
        elif json_data[i]['type'] == 'hostname':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                if ':' in kwargs['var_dict'][i]: validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE): validating.dns_name_ws(i, kwargs)
                else: validating.ip_address_ws(i, kwargs)
        elif json_data[i]['type'] == 'list_of_email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', kwargs)
        elif json_data[i]['type'] == 'email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.email_ws(i, kwargs)
        elif json_data[i]['type'] == 'integer':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.number_check(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_domains':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for domain in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'domain_{count}'] = domain
                    validating.domain_ws(f'domain_{count}', kwargs)
                    kwargs['var_dict'].pop(f'domain_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_hosts':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = hostname
                    if ':' in hostname: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE): validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_integer':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.number_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.string_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_values':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.list_values(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_vlans':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.vlans(i, kwargs)
        elif json_data[i]['type'] == 'string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.string_pattern(i, json_data, kwargs)
        else: pcolor.Red(f"error validating.  Type not found {json_data[i]['type']}. 2."); len(False); sys.exit(1)
    for i in json_data['optional_args']:
        if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
            if re.search(r'^module_[\d]+$', i): validating.list_values_key('modules', i, json_data, kwargs)
            elif json_data[i]['type'] == 'boolean': validating.boolean(i, json_data, kwargs)
            elif json_data[i]['type'] == 'domain': validating.domain_ws(i, kwargs)
            elif json_data[i]['type'] == 'list_of_email':
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', json_data, kwargs)
            elif json_data[i]['type'] == 'email': validating.email_ws(i, json_data, kwargs)
            elif json_data[i]['type'] == 'hostname':
                if ':' in kwargs['var_dict'][i]: validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE): validating.dns_name_ws(i, kwargs)
                else: validating.ip_address_ws(i, kwargs)
            elif json_data[i]['type'] == 'integer': validating.number_check(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_integer': validating.number_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_hosts':
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE): validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_macs':
                count = 1
                for mac in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = mac
                    validating.mac_address(f'{i}_{count}', kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_string': validating.string_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_values': validating.list_values(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_vlans':  validating.vlans(i, kwargs)
            elif json_data[i]['type'] == 'mac_address':    validating.mac_address(i, kwargs)
            elif json_data[i]['type'] == 'string':         validating.string_pattern(i, json_data, kwargs)
            else: pcolor.Red(f"error validating.  Type not found {json_data[i]['type']}. 3."); len(False); sys.exit(1)
    return kwargs

#=============================================================================
# Function - Check VLAN exists in VLAN Policy
#=============================================================================
def validate_vlan_in_policy(vlan_policy_list, vlan_id):
    valid = False
    while valid == False:
        vlan_count = 0
        for vlan in vlan_policy_list:
            if int(vlan_id) == 1: vlan_count = 1; continue
            if int(vlan) == int(vlan_id): vlan_count = 1; continue
        if vlan_count == 1: valid = True; return valid
        else:
            pcolor.Red(f'\n-------------------------------------------------------------------------------------------\n')
            pcolor.Red(f'  VLAN {vlan_id} not found in the VLAN Policy List.  Please us a VLAN from the list below:')
            pcolor.Red(f'  {vlan_policy_list}')
            pcolor.Red(f'\n-------------------------------------------------------------------------------------------\n')
            return valid

#=============================================================================
# Function - Validate IPMI Key value
#=============================================================================
def validate_ipmi_key(varValue):
    valid_count = 0
    varValue = varValue.capitalize()
    if ((varValue < '0' or varValue > '9') and (varValue < 'A' or varValue > 'F')): valid_count += 1
    if not validators.length(varValue, min=2, max=40): valid_count += 1
    if not len(varValue) % 2 == 0: valid_count += 1
    if not valid_count == 0:
        pcolor.Red(f'\n{"-"*108}\n')
        pcolor.Red(f'   Error with ipmi_key!!  The encryption key should have an even number of ')
        pcolor.Red(f'   hexadecimal characters and not exceed 40 characters.\n')
        pcolor.Red(f'   Valid Hex Characters are:')
        pcolor.Red(f'    - {string.hexdigits}')
        pcolor.Red(f'\n{"-"*108}\n')
        return False
    else: return True

#=============================================================================
# Function - Validate Sensitive Strings
#=============================================================================
def validate_sensitive(secure_value, kwargs):
    invalid_count = 0
    if not validators.length(secure_value, min=int(kwargs.jdata.minLength), max=int(kwargs.jdata.maxLength)):
        invalid_count += 1
        pcolor.Red(f'\n{"-"*108}\n')
        pcolor.Red(f'   !!! {kwargs.sensitive_var} is Invalid!!!')
        pcolor.Red(f'   Length Must be between {kwargs.jdata.minLength} and {kwargs.jdata.maxLength} characters.')
        pcolor.Red(f'\n{"-"*108}\n')
    if not re.search(kwargs.jdata.pattern, secure_value):
        invalid_count += 1
        pcolor.Red(f'\n{"-"*108}\n')
        pcolor.Red(f'   !!! Invalid Characters in {kwargs.sensitive_var}.  The allowed characters are:')
        pcolor.Red(f'   - "{kwargs.jdata.pattern}"')
        pcolor.Red(f'\n{"-"*108}\n')
    if invalid_count == 0: return True
    else: return False

#=============================================================================
# Function - Validate Sensitive Strings
#=============================================================================
def validate_strong_password(secure_value, kwargs):
    invalid_count = 0; valid_count = 0
    if re.search(kwargs.username, secure_value, re.IGNORECASE): invalid_count += 1
    if not validators.length(str(secure_value), min=int(kwargs.jdata.minLength), max=int(kwargs.jdata.maxLength)):
        invalid_count += 1
    else: valid_count +=1
    if re.search(r'[a-z]', secure_value): valid_count += 1
    if re.search(r'[A-Z]', secure_value): valid_count += 1
    if re.search(r'[0-9]', secure_value): valid_count += 1
    if re.search(r'[\!\@\#\$\%\^\&\*\-\_\+\=]', secure_value): valid_count += 1
    if not invalid_count == 0 and valid_count >= 4:
        pcolor.Red(f'\n{"-"*108}\n')
        pcolor.Red(f"   Error with {kwargs.sensitive_var}! The password failed one of the following complexity rules:")
        pcolor.Red(f'     - The password must have a minimum of 8 and a maximum of 20 characters.')
        pcolor.Red(f"     - The password must not contain the User's Name.")
        pcolor.Red(f'     - The password must contain characters from three of the following four categories.')
        pcolor.Red(f'       * English uppercase characters (A through Z).')
        pcolor.Red(f'       * English lowercase characters (a through z).')
        pcolor.Red(f'       * Base 10 digits (0 through 9).')
        pcolor.Red(f'       * Non-alphabetic characters (! , @, #, $, %, ^, &, *, -, _, +, =)')
        pcolor.Red(f'\n{"-"*108}\n')
        return False
    else: return True

#=============================================================================
# Function - Prompt for Answer to Question from List
#=============================================================================
def variable_from_list(kwargs):
    #=========================================================================
    # Set Function Variables
    #=========================================================================
    default     = kwargs.jdata.default
    description = kwargs.jdata.description
    optional    = False
    title       = kwargs.jdata.title
    if not kwargs.jdata.get('multi_select'): kwargs.jdata.multi_select = False
    #=========================================================================
    # Sort the Variables
    #=========================================================================
    if kwargs.jdata.get('sort') == False: vars = kwargs.jdata.enum
    else: vars = sorted(kwargs.jdata.enum, key=str.casefold)
    valid = False
    while valid == False:
        pcolor.LightPurple(f'\n{"-"*108}\n')
        if '\n' in description:
            description = description.split('\n')
            for line in description: pcolor.LightGray(line)
        else: pcolor.LightGray(description)
        if kwargs.jdata.get('multi_select') == True:
            pcolor.Yellow('\n     Note: Answer can be:\n       * Single: 1\n       * Multiple: `1,2,3` or `1-3,5-6`')
        if kwargs.jdata.get('multi_select') == True: pcolor.Yellow(f'    Select Option(s) Below:')
        else: pcolor.Yellow(f'\n    Select an Option Below:')
        for index, value in enumerate(vars):
            index += 1
            if value == default: default_index = index
            if   index < 10:  pcolor.Cyan(f'      {index}. {value}')
            elif index < 100: pcolor.Cyan(f'     {index}. {value}')
            elif index > 99:  pcolor.Cyan(f'    {index}. {value}')
        if kwargs.jdata.get('multi_select') == True:
            if kwargs.jdata.get('optional') == True:
                optional = True
                var_selection   = input(f'\nPlease Enter the Option Number(s) to select for {title}.  [press enter to skip]: ')
            elif not default == '':
                var_selection   = input(f'\nPlease Enter the Option Number(s) to select for {title}.  [{default_index}]: ')
            else: var_selection = input(f'\nPlease Enter the Option Number(s) to select for {title}: ')
        else:
            if kwargs.jdata.get('optional') == True:
                optional = True
                var_selection   = input(f'\nPlease Enter the Option Number to select for {title}.  [press enter to skip]: ')
            elif not default == '':
                var_selection   = input(f'\nPlease Enter the Option Number to select for {title}.  [{default_index}]: ')
            else: var_selection = input(f'\nPlease Enter the Option Number to select for {title}: ')
        if   kwargs.jdata.get('optional') == True and var_selection == '' and kwargs.jdata.multi_select == False: return '', True
        elif kwargs.jdata.get('optional') == True and var_selection == '' and kwargs.jdata.multi_select == True:  return [], True
        elif not default == '' and var_selection == '':
            var_selection = default_index
        if kwargs.jdata.multi_select == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(vars):
                index += 1
                if int(var_selection) == index: selection = value; valid = True
        elif kwargs.jdata.multi_select == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index: var_count += 1; selection.append(value)
            if var_count == var_length: valid = True
            else: pcolor.Red(f'\n{"-"*108}\n\n  The list of Vars {var_list} did not match the available list.\n\n{"-"*108}\n')
        if valid == False: message_invalid_selection()
    return selection, valid

#=============================================================================
# Function - Prompt User for Answer to Question
#=============================================================================
def variable_prompt(kwargs):
    #=========================================================================
    # Improper Value Notifications
    #=========================================================================
    def invalid_boolean(title, answer):
        pcolor.Red(f'\n{"-"*108}\n   `{title}` value of `{answer}` is Invalid!!! Please enter `Y` or `N`.\n{"-"*108}\n')
    def invalid_integer(title, answer):
        pcolor.Red(f'\n{"-"*108}\n   `{title}` value of `{answer}` is Invalid!!!  Valid range is `{minimum}-{maximum}`.\n{"-"*108}\n')
    def invalid_string(title, answer):
        pcolor.Red(f'\n{"-"*108}\n   `{title}` value of `{answer}` is Invalid!!!\n{"-"*108}\n')
    #=========================================================================
    # Set Function Variables
    #=========================================================================
    default     = kwargs.jdata.default
    description = kwargs.jdata.description
    optional    = False
    title       = kwargs.jdata.title
    #=========================================================================
    # Print `description` if not enum
    #=========================================================================
    if not kwargs.jdata.get('enum'): pcolor.LightPurple(f'\n{"-"*108}\n'); pcolor.LightGray(f'{description}\n')
    #=========================================================================
    # Prompt User for Answer
    #=========================================================================
    valid = False
    while valid == False:
        if kwargs.jdata.get('enum'):  answer, valid = variable_from_list(kwargs)
        elif kwargs.jdata.type == 'boolean':
            if default == True: default = 'Y'
            else: default = 'N'
            answer = input(f'\nEnter `Y` for `True` or `N` for `False` for `{title}`. [{default}]: ')
            if answer == '':
                if default == 'Y': answer = True
                elif default == 'N': answer = False
                valid = True
            elif answer == 'N': answer = False; valid = True
            elif answer == 'Y': answer = True;  valid = True
            else: invalid_boolean(title, answer)
        elif kwargs.jdata.type == 'integer':
            maximum = kwargs.jdata.maximum
            minimum = kwargs.jdata.minimum
            if kwargs.jdata.get('optional') == True:
                optional = True
                answer = input(f'Enter the value for {title} [press enter to skip]: ')
            else: answer = input(f'Enter the Value for {title}. [{default}]: ')
            if optional == True and answer == '': valid = True
            elif answer == '': answer = default
            if optional == False:
                if re.fullmatch(r'^[0-9]+$', str(answer)):
                    if kwargs.jdata.title == 'snmp_port':
                        valid = validating.snmp_port(title, answer, minimum, maximum)
                    else: valid = validating.number_in_range(title, answer, minimum, maximum)
                else: invalid_integer(title, answer)
        elif kwargs.jdata.type == 'string':
            if kwargs.jdata.get('optional') == True:
                optional = True
                answer = input(f'Enter the value for {title} [press enter to skip]: ')
            elif not default == '': answer = input(f'Enter the value for {title} [{default}]: ')
            else: answer = input(f'Enter the value for {title}: ')
            if optional == True and answer == '': valid = True
            elif answer == '': answer = default; valid = True
            elif not answer == '':
                maxLength = kwargs.jdata.maxLength
                minLength = kwargs.jdata.minLength
                pattern   = kwargs.jdata.pattern
                valid = validating.length_and_regex(answer, minLength, maxLength, pattern, title)
        else: invalid_string(title, answer)
    if kwargs.jdata.get('optional'): kwargs.jdata.pop('optional')
    if kwargs.jdata.get('multi_select'): kwargs.jdata.pop('multi_select')
    return answer

#=============================================================================
# Function - Collapse VLAN List
#=============================================================================
def vlan_list_format(vlan_list_expanded):
    vlan_list = sorted(vlan_list_expanded)
    vgroups   = itertools.groupby(vlan_list, key=lambda item, c=itertools.count():item-next(c))
    tempvlans = [list(g) for k, g in vgroups]
    vlanList  = [str(x[0]) if len(x) == 1 else f'{x[0]}-{x[-1]}' for x in tempvlans]
    vlan_list = ','.join(vlanList)
    return vlan_list

#=============================================================================
# Function - Expand VLAN List
#=============================================================================
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.search(r'-', v):
                a,b = v.split('-'); a = int(a); b = int(b); vrange = range(a,b+1)
                for vl in vrange: full_vlan_list.append(int(vl))
            else: full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-'); a = int(a); b = int(b); vrange = range(a,b+1)
        for v in vrange: full_vlan_list.append(int(v))
    else: full_vlan_list.append(int(vlan_list))
    return full_vlan_list

#=============================================================================
# Function - To Request Native VLAN
#=============================================================================
def vlan_native_function(vlan_policy_list, vlan_list):
    native_count = 0
    nativeVlan = ''
    nativeValid = False
    while nativeValid == False:
        nativeVlan = input('Do you want to Configure one of the VLANs as a Native VLAN?  [press enter to skip]:')
        if nativeVlan == '': nativeValid = True
        else:
            for vlan in vlan_policy_list:
                if int(nativeVlan) == int(vlan): native_count = 1; break
            if not native_count == 1: message_invalid_native_vlan(nativeVlan, vlan_list)
            else: nativeValid = True
    return nativeVlan

#=============================================================================
# Function - Prompt for VLANs and Configure Policy
#=============================================================================
def vlan_pool(name):
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n{"-"*108}\n')
        pcolor.Cyan(f'  The allowed vlan list can be in the format of:')
        pcolor.Cyan(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
        pcolor.Cyan(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*108}\n')
        VlanList = input(f'Enter the VLAN or List of VLANs to assign to the Domain VLAN Pool {name}: ')
        if not VlanList == '':
            vlanListExpanded = vlan_list_full(VlanList)
            valid_vlan = True
            for vlan in vlanListExpanded:
                valid_vlan = validating.number_in_range('VLAN ID', vlan, 1, 4094)
                if valid_vlan == False: break
            if valid_vlan == False:
                pcolor.Red(f'\n{"-"*108}\n')
                pcolor.Red(f'  !!!Error!!!\n  With VLAN(s) assignment. VLAN List: "{VlanList}" is not Valid.')
                pcolor.Red(f'  The allowed vlan list can be in the format of:')
                pcolor.Red(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
                pcolor.Red(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*108}\n')
            else: valid = True
        else:
            pcolor.Red(f'\n{"-"*108}\n')
            pcolor.Red(f'  The allowed vlan list can be in the format of:')
            pcolor.Red(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
            pcolor.Red(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*108}\n')
    return VlanList,vlanListExpanded

#=============================================================================
# Function - Obtain Windows Language Dictionary
#=============================================================================
def windows_languages(windows_language, kwargs):
    kwargs.windows_languages = json.load(open(os.path.join(kwargs.script_path, 'variables', 'windowsLocals.json'), 'r'))
    language = [e for e in kwargs.windows_languages if (
        (DotMap(e)).language.replace('(', '_')).replace(')', '_') == (windows_language.language_pack.replace('(', '_')).replace(')', '_')]
    if len(language) == 1: language = DotMap(language[0])
    else:
        pcolor.Red(f'Failed to Map `{windows_language.language_pack}` to a Windows Language.')
        pcolor.Red(f'Available Languages are:')
        for e in kwargs.windows_languages: pcolor.Red(f'  * {(DotMap(e)).language}')
        len(False); sys.exit(1)
    kwargs.language = DotMap(
        ui_language        = language.code,
        input_locale       = (re.search('\\((.*)\\)', language.local)).group(1),
        layered_driver     = windows_language.layered_driver,
        secondary_language = '')
    if language.get('secondary_language'):
        if type(language.secondary_language) == list:
            kwargs.language.secondary_language   = language.secondary_language[0]
        else: kwargs.language.secondary_language = language.secondary_language
    if kwargs.language.layered_driver == 0: kwargs.language.layered_driver = ''
    return kwargs

#=============================================================================
# Function - Obtain Windows Timezone
#=============================================================================
def windows_timezones(kwargs):
    kwargs.windows_timezones = DotMap(json.load(open(os.path.join(kwargs.script_path, 'variables', 'windowsTimeZones.json'), 'r')))
    tz = pytz.timezone(kwargs.timezone)
    june = pytz.utc.localize(datetime(2023, 6, 2, 12, 1, tzinfo=None))
    december = pytz.utc.localize(datetime(2023, 12, 2, 12, 1, tzinfo=None))
    june_dst = june.astimezone(tz).dst() != timedelta(0)
    dec_dst  = december.astimezone(tz).dst() != timedelta(0)
    if june_dst == True and dec_dst == False: kwargs.disable_daylight = 'false'
    elif june_dst == False and dec_dst == True: kwargs.disable_daylight = 'false'
    elif june_dst == False and dec_dst == False: kwargs.disable_daylight = 'true'
    else:
        pcolor.Red(f'unknown Timezone Result for {kwargs.timezone}')
        len(False); sys.exit(1)
    windows_timezone = [k for k, v in kwargs.windows_timezones.items() if v == kwargs.timezone]
    if len(windows_timezone) == 1: kwargs.windows_timezone = windows_timezone[0]
    else:
        pcolor.Red(f'Failed to Map `{kwargs.timezone}` to a Windows Timezone.')
        pcolor.Red(f'Available Languages are:')
        for k,v in kwargs.windows_timezones.items(): pcolor.Red(f'  * {k}: {v}')
        len(False); sys.exit(1)
    return kwargs

#=============================================================================
# Function to Determine which sites to write files to.
#=============================================================================
def write_to_repo_folder(pol_vars, kwargs):
    baseRepo   = kwargs.args.dir
    dest_file  = kwargs.dest_file
    #=========================================================================
    # Setup jinja2 Environment
    #=========================================================================
    template_path = pkg_resources.resource_filename(f'policies', 'templates/')
    templateLoader = jinja2.FileSystemLoader(searchpath=(template_path + 'provider/'))
    templateEnv = jinja2.Environment(loader=templateLoader)
    #=========================================================================
    # Define the Template Source
    #=========================================================================
    template = templateEnv.get_template(kwargs.template_file)
    #=========================================================================
    # Make sure the Destination Path and Folder Exist
    #=========================================================================
    if not os.path.isdir(os.path.join(baseRepo)): dest_path = f'{os.path.join(baseRepo)}'; os.makedirs(dest_path)
    dest_dir = os.path.join(baseRepo)
    if not os.path.exists(os.path.join(dest_dir, dest_file)):
        create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'; os.system(create_file)
    tf_file = os.path.join(dest_dir, dest_file)
    wr_file = open(tf_file, 'w')
    #=========================================================================
    # Render Payload and Write to File
    #=========================================================================
    pol_vars = json.loads(json.dumps(pol_vars))
    pol_vars = {'keys':pol_vars}
    payload = template.render(pol_vars)
    wr_file.write(payload)
    wr_file.close()
