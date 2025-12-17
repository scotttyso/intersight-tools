#=============================================================================
# Source Modules
#=============================================================================
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
import sys
try:
    from classes         import ezfunctions, isight, pcolor, questions
    from datetime        import datetime
    from dotmap          import DotMap
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
    from stringcase      import snakecase
    import dateutil.relativedelta, json, numpy, pytz, openpyxl, os, re, shutil, urllib3, yaml
except ImportError as e:
    prRed(f'classes/day2tools.py line 6 - !!! ERROR !!!\n{e.__class__.__name__}')
    prRed(f" Module {e.name} is required to run this script")
    prRed(f" Install the module using the following: `pip install {e.name}`")
    sys.exit(1)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class tools(object):
    def __init__(self, type):
        self.type = type

    #=========================================================================
    # Function - Add Policies
    #=========================================================================
    def add_policies(self, kwargs):
        #=====================================================================
        # Prompt User for Source and Destination Organization
        #=====================================================================
        orgs = list(kwargs.org_moids.keys())
        kwargs.jdata             = kwargs.ezwizard.setup.properties.organization
        kwargs.jdata.description = 'Select the Source Organization for the pools/policies:'
        kwargs.jdata.enum        = orgs
        source_org               = ezfunctions.variable_prompt(kwargs)
        kwargs.jdata.description = 'Select the Organization for the Profiles:'
        destination_org          = ezfunctions.variable_prompt(kwargs)
        #=====================================================================
        # Prompt for Profile Type and obtain Pools/Policies
        #=====================================================================
        kwargs     = questions.profiles.profile_type(kwargs)
        kwargs     = questions.policies.build_policy_list(kwargs)
        pool_types = []
        if kwargs.target_platform == 'FIAttached':
            kwargs.jdata       = kwargs.ezwizard.setup.properties.server_type
            kwargs.server_type = ezfunctions.variable_prompt(kwargs)
            for e in ['ip', 'iqn', 'mac', 'wwnn', 'wwpn']:
                if e in kwargs.pool_list: kwargs.pool_list.remove(e)
            kwargs.jdata      = kwargs.ezwizard.setup.properties.pool_types
            kwargs.jdata.enum = kwargs.pool_list
            pool_types        = ezfunctions.variable_prompt(kwargs)
        for e in ['ethernet_adapter', 'ethernet_network_control', 'ethernet_network_group', 'ethernet_qos', 'fc_zone', 'fibre_channel_adapter',
                  'fibre_channel_network', 'fibre_channel_qos', 'firmware_authenticate', 'iscsi_adapter', 'iscsi_boot', 'iscsi_static_target']:
            if e in kwargs.policy_list: kwargs.policy_list.remove(e)
        kwargs.jdata      = kwargs.ezwizard.setup.properties.policy_types
        kwargs.jdata.enum = kwargs.policy_list
        policy_types      = ezfunctions.variable_prompt(kwargs)
        #=====================================================================
        # Prompt User for Policies to Attach
        #=====================================================================
        pool_types.extend(policy_types)
        update_types = pool_types
        kwargs.org   = source_org
        kwargs.policy_bucket = []
        kwargs.pools         = []
        accept_policies = False
        while accept_policies == False:
            policy_names = DotMap()
            for e in update_types:
                api_filter = f"Organization.Moid eq '{kwargs.org_moids[source_org].moid}'"
                kwargs = kwargs | DotMap(api_filter = api_filter, method = 'get', uri = kwargs.ezdata[e].intersight_uri)
                kwargs = isight.api(e).calls(kwargs)
                kwargs[f'{e}_results'] = kwargs.results
                for d in kwargs[f'{e}_results']: kwargs[e][d.Name] = d
                policies = sorted([d.Name for d in kwargs[f'{e}_results']])
                kwargs.jdata  = DotMap(
                    default     = '',
                    description = f'Select the {e} policy from source org: `{source_org}` to attach to {kwargs.target_platform}(s) in org: `{destination_org}`.',
                    enum        = policies, title = f'{e} Policy', type = 'string')
                if type(kwargs.server_type) == str:
                    kwargs.jdata.description = kwargs.jdata.description.replace(kwargs.target_platform, kwargs.server_type)
                if re.search('^resource|uuid$', e):
                    kwargs.jdata.description.replace('policy', 'pool')
                    kwargs.jdata.title.replace('Policy', 'Pool')
                policy_name = ezfunctions.variable_prompt(kwargs)
                policy_names[e] = policy_name
                indx = next((index for (index, d) in enumerate(kwargs[f'{e}_results']) if d['Name'] == policy_name), None)
                if re.search('^resource|uuid$', e):
                    kwargs.pools.append(DotMap(Moid=kwargs[f'{e}_results'][indx].Moid,ObjectType=kwargs[f'{e}_results'][indx].ObjectType))
                else: kwargs.policy_bucket.append(DotMap(Moid=kwargs[f'{e}_results'][indx].Moid,ObjectType=kwargs[f'{e}_results'][indx].ObjectType))
            answer = questions.prompt_user.to_accept('the `policy bucket update`', policy_names, kwargs)
            if answer == True: accept_policies = True
        #=====================================================================
        # Update the Profiles
        #=====================================================================
        kwargs.org = destination_org
        if len(kwargs.policy_bucket) > 0: kwargs = tools(kwargs.target_platform).profiles_update(kwargs)

    #=========================================================================
    # Function - Add Vlans
    #=========================================================================
    def add_vlans(self, kwargs):
        #=====================================================================
        # Validate YAML configuration file is defined.
        #=====================================================================
        if kwargs.args.yaml_file != None:
            kwargs.ydata = DotMap(yaml.safe_load(open(kwargs.args.yaml_file, 'r'))).add_vlans
        else:
            prRed(f'\n{"-"*108}\n\n  Missing Required YAML File Argument `-y`.  Exiting Process.')
            prRed(f'\n{"-"*108}\n')
            len(False); sys.exit(1)
        #=====================================================================
        # Get VLAN List and Organizations from YAML configuration file.
        #=====================================================================
        for org in kwargs.ydata.organizations:
            #=================================================================
            # Query the API for the VLAN Policies
            #=================================================================
            kwargs.org = org
            kwargs     = isight.api_get(False, [e.name for e in kwargs.ydata.vlan], 'vlan', kwargs)
            kwargs.vlan_policy_moids = kwargs.pmoids
            for vlan_policy in kwargs.ydata.vlan:
                pvars  = DotMap(name = vlan_policy.name, vlans = [])
                kwargs = tools('add_vlans').add_vlans_to_vlan_policies(vlan_policy, pvars, kwargs)
                kwargs.class_path = 'policies,vlan'
                kwargs = ezfunctions.ez_append(pvars, kwargs)
                #=============================================================
                # Query the API for the Ethernet Network Group Policies
                #=============================================================
                if len(kwargs.ydata.ethernet_network_group) > 0:
                    kwargs = tools('eng').add_vlans_update_ethernet_network_groups(kwargs)
                #=================================================================
                # Create LAN Connectivity Policies if they are defined
                #=================================================================
                if len(kwargs.ydata.lan_connectivity_templates) > 0:
                    kwargs = tools('lan_connectivity').add_vlans_lan_connectivity_templates(vlan_policy, kwargs)
        #=====================================================================
        # Deploy Policies
        #=====================================================================
        kwargs = isight.imm.deploy(kwargs)
        shutil.rmtree('Intersight')

    #=========================================================================
    # Function - Add Vlans - LAN Connectivity Policies
    #=========================================================================
    def add_vlans_lan_connectivity_templates(self, vlan_policy, kwargs):
        for e in vlan_policy.vlans:
            #print(e)
            #=================================================================
            # Build Ethernet Network Group Dictionaries
            #=================================================================
            for lcp in e.lan_connectivity_templates:
                for v in lcp.vnics:
                    for eng in v.ethernet_network_group_policies:
                        if '{{vlan_id}}' in eng:
                            pvars = DotMap(allowed_vlans = str(e.vlan_id), name = eng.replace('{{vlan_id}}', str(e.vlan_id)), native_vlan = e.vlan_id)
                            kwargs.class_path = 'policies,ethernet_network_group'
                            kwargs            = ezfunctions.ez_append(pvars, kwargs)
            #=================================================================
            # Build LAN Connectivity Dictionaries
            #=================================================================
            for lcp in e.lan_connectivity_templates:
                indx = next((index for (index, d) in enumerate(kwargs.ydata.lan_connectivity_templates) if d.policy_reference == lcp.policy_reference), None)
                if indx != None:
                    template = kwargs.ydata.lan_connectivity_templates[indx]
                    tkeys    = list(template.keys())
                    if 'target_platform' in tkeys: platform = template.target_platform
                    else: platform = 'FIAttached'
                    pvars = DotMap(name = template.name.replace('{{vlan_id}}', str(e.vlan_id)), target_platform = platform, vnics = [])
                    for x in range(0,len(template.vnics)):
                        pdict = DotMap()
                        plist = list(kwargs.ezdata['lan_connectivity.vnics'].properties.keys())
                        vnic  = template.vnics[x]
                        vkeys = list(vnic.keys())
                        for p in plist:
                            if   p == 'ethernet_network_group_policies': pdict[p] = [d.replace('{{vlan_id}}', str(e.vlan_id)) for d in lcp.vnics[x][p]]
                            elif p == 'ethernet_adapter_policy': pdict[p] = lcp.vnics[x][p]
                            elif p in vkeys: pdict[p] = vnic[p]
                        pvars.vnics.append(pdict)
                        kwargs.class_path = 'policies,lan_connectivity'
                        kwargs            = ezfunctions.ez_append(pvars, kwargs)
        #=====================================================================
        # Return kwargs
        #=====================================================================
        return kwargs

    #=========================================================================
    # Function - Add Vlans - Update Ethernet Network Groups
    #=========================================================================
    def add_vlans_update_ethernet_network_groups(self, kwargs):
        #=====================================================================
        # Query API for Ethernet Network Group Policies
        #=====================================================================
        kwargs = isight.api_get(False, kwargs.ydata.ethernet_network_group, 'ethernet_network_group', kwargs)
        eng_results = kwargs.results
        for e in kwargs.ydata.ethernet_network_group:
            indx              = next((index for (index, d) in enumerate(eng_results) if d['Name'] == e), None)
            allowed_vlans     = ezfunctions.vlan_list_full(eng_results[indx].VlanSettings.AllowedVlans)
            allowed_vlans.extend(kwargs.add_vlans)
            allowed_vlans     = ezfunctions.vlan_list_format(list(numpy.unique(numpy.array(allowed_vlans))))
            pvars             = DotMap(allowed_vlans = allowed_vlans, name = e)
            kwargs.class_path = 'policies,ethernet_network_group'
            kwargs            = ezfunctions.ez_append(pvars, kwargs)
        return kwargs

    #=========================================================================
    # Function - Add Vlans - Add VLANs to VLAN Policies
    #=========================================================================
    def add_vlans_to_vlan_policies(self, i, pvars, kwargs):
        #=====================================================================
        # Get VLAN Policy & VLAN(s) attributes
        #=====================================================================
        vlan_policy_moid  = kwargs.vlan_policy_moids[pvars.name].moid
        kwargs = kwargs | DotMap(api_filter = f"EthNetworkPolicy.Moid eq '{vlan_policy_moid}'", method = 'get', uri = kwargs.ezdata['vlan.vlans'].intersight_uri)
        kwargs = isight.api('vlan.vlans').calls(kwargs)
        kwargs = kwargs | DotMap(method = 'get_by_moid', pmoid = kwargs.results[-1].MulticastPolicy.Moid, uri = kwargs.ezdata.multicast.intersight_uri)
        kwargs = isight.api('multicast').calls(kwargs)
        multicast_name   = (f'{kwargs.org_names[kwargs.results.Organization.Moid]}/{kwargs.results.Name}').replace(f'{kwargs.org}/', '')
        kwargs.add_vlans = []
        #=====================================================================
        # Update VLAN Policy with VLAN List
        #=====================================================================
        for e in i.vlans:
            kwargs.add_vlans.append(e.vlan_id)
            pvars.vlans.append(DotMap(multicast_policy = multicast_name, name = e.name, vlan_list = str(e.vlan_id)))
        return kwargs

    #=========================================================================
    # Function - Clone Policies
    #=========================================================================
    def audit_logs(self, kwargs):
        #=====================================================================
        # Prompt User for Source and Destination Organization
        #=====================================================================
        kwargs.jdata = DotMap(default = 1, minimum = 1, maximum = 6, title = 'Audit Record Months', type='integer',
                              description = 'Enter the Number of Months, between 1 and 6, to obtain Audit Records for.')
        months = ezfunctions.variable_prompt(kwargs)
        today      = datetime.now()
        last_month = today + dateutil.relativedelta.relativedelta(months=-int(months))
        api_filter = f"CreateTime gt {last_month.strftime('%Y-%m-%d')}T00:00:00.000Z and CreateTime lt {today.strftime('%Y-%m-%d')}T23:59:00.000Z"
        kwargs = kwargs | DotMap(api_filter = api_filter, method = 'get', org = 'default', uri = 'aaa/AuditRecords')
        kwargs = isight.api('audit_records').calls(kwargs)
        audit_dict = DotMap(); ucount = DotMap(); users = []
        for e in kwargs.results:
            day = e.CreateTime.split('T')[0]
            if not audit_dict.get(day): audit_dict[day]
            if not audit_dict[day].get(e.Email): audit_dict[day][e.Email].count = 1
            else: audit_dict[day][e.Email].count += 1
        with open('audit_log.json', 'w') as json_file:
            json.dump(audit_dict, json_file, indent=4)
        for k,v in audit_dict.items(): users.extend(list(v.keys()))
        unique_users = list(filter(None, sorted(list(numpy.unique(numpy.array(users))))))
        for e in unique_users: ucount[e].occurances = users.count(e)
        pcolor.Cyan(f'\n{"-"*108}\n')
        pcolor.Yellow(json.dumps(ucount, indent=4))
        pcolor.Cyan(f'\n{"-"*54}\n')
        pcolor.Cyan(f'  Saved Full Audit Data to `{os.getcwd()}/audit_log.json`')
        pcolor.Cyan(f'\n{"-"*108}')
        

    #=========================================================================
    # Function - Clone Policies
    #=========================================================================
    def clone_policies(self, kwargs):
        #=====================================================================
        # Prompt User for Source and Destination Organization
        #=====================================================================
        orgs = list(kwargs.org_moids.keys())
        kwargs.jdata             = kwargs.ezwizard.setup.properties.organization
        kwargs.jdata.description = 'Select the Source Organization to clone the pools/policies from:'
        kwargs.jdata.enum        = orgs
        source_org               = ezfunctions.variable_prompt(kwargs)
        kwargs.jdata.description = 'Select the Destination Organization to clone the policies to:'
        destination_org          = ezfunctions.variable_prompt(kwargs)
        #=====================================================================
        # Prompt for Profile Type and obtain Pools/Policies
        #=====================================================================
        kwargs = questions.profiles.profile_type(kwargs)
        kwargs = questions.policies.build_policy_list(kwargs)
        for e in ['imc_access', 'iscsi_boot', 'firmware_authenticate', 'lan_connectivity', 'ldap', 'local_user', 'port', 'san_connectivity', 'storage', 'vlan', 'vsan']:
            if e in kwargs.policy_list: kwargs.policy_list.remove(e)
        kwargs.jdata      = kwargs.ezwizard.setup.properties.policy_types
        kwargs.jdata.enum = kwargs.policy_list,
        policy_types      = ezfunctions.variable_prompt(kwargs)
        pool_types        = []
        if kwargs.target_platform == 'FIAttached':
            kwargs.jdata      = kwargs.ezwizard.setup.properties.pool_types
            kwargs.jdata.enum = kwargs.pool_list,
            pool_types        = ezfunctions.variable_prompt(kwargs)
        #=====================================================================
        # Prompt User for Policies to Clone
        #=====================================================================
        pool_types.extend(policy_types)
        clone_types = pool_types
        kwargs.org  = source_org
        for e in clone_types:
            api_filter = f"Organization.Moid eq '{kwargs.org_moids[source_org].moid}'"
            kwargs = kwargs | DotMap(api_filter = api_filter, method = 'get', uri = kwargs.ezdata[e].intersight_uri)
            kwargs = isight.api(e).calls(kwargs)
            kwargs[f'{e}_results'] = kwargs.results
            for d in kwargs[f'{e}_results']: kwargs[e][d.Name] = d
            policies = sorted([d.Name for d in kwargs[f'{e}_results']])
            kwargs.jdata  = DotMap(
                default       = '',
                description   = f'Select the `{e}` policies to clone from source org: `{source_org}` to destination org: `{destination_org}`.',
                multi_select  = True,
                enum          = policies, title = f'{e} Policies', type = 'string')
            if re.search('^ip|iqn|mac|resource|uuid|wwnn|wwpn$', e):
                kwargs.jdata.description.replace('policies', 'pools')
                kwargs.jdata.title.replace('Policies', 'Pools')
            clone_policies = ezfunctions.variable_prompt(kwargs)
            key_list = ['Description', 'Name', 'Tags']
            for k,v in kwargs.ezdata[e].allOf[1].properties.items():
                if v.type == 'array': key_list.append(v['items'].intersight_api)
                elif v.type == 'object': key_list.append(v.intersight_api)
                elif re.search('boolean|integer|string', v.type):
                    if re.search(r'\$ref\:', v.intersight_api): key_list.append(v.intersight_api.split(':')[1])
                    else: key_list.append(v.intersight_api)
                else: pcolor.Yellow(json.dumps(v, indent=4)); sys.exit(1)
            kwargs.bulk_list = []
            key_list = list(numpy.unique(numpy.array(key_list)))
            for d in clone_policies:
                api_body = kwargs[e][d]
                for key in list(api_body.keys()):
                    if not key in key_list: api_body.pop(key)
                api_body.ObjectType   = kwargs.ezdata[e].object_type
                api_body.Organization = dict(Moid = kwargs.org_moids[destination_org].moid, ObjectType = 'organization.Organization')
                kwargs.bulk_list.append(api_body.toDict())
            #=================================================================
            # POST Bulk Request if Post List > 0
            #=================================================================
            if len(kwargs.bulk_list) > 0:
                kwargs.uri = kwargs.ezdata[e].intersight_uri
                kwargs     = isight.imm(e).bulk_request(kwargs)

    #=========================================================================
    # Function - HCL Status
    #=========================================================================
    def hcl_status(self, kwargs):
        kwargs.org = 'default'
        pdict = DotMap()
        # Obtain Server Profile Data
        for e in kwargs.json_data:
            if 'Cisco' in e.Hostname.Manufacturer:
                pdict[e.Serial] = DotMap(
                    domain = 'Standalone', model = e.Hostname.Model, serial = e.Serial, server_dn = 'unknown', server_profile = 'unassigned',
                    firmware = 'unknown', vcenter = e.vCenter, cluster = e.Cluster, hostname = e.Hostname.Name, version = e.Hostname.Version,
                    build = e.Hostname.Build, hcl_check = 'unknown', hcl_status = 'unknown', hcl_reason = 'unknown', toolName = e.Name,
                    toolDate = e.InstallDate, toolVer = e.Version, moid = 'unknown'
                )
        kwargs = kwargs | DotMap(method = 'get', names = list(pdict.keys()), uri = 'compute/PhysicalSummaries')
        kwargs = isight.api('serial_number').calls(kwargs)
        registered_devices  = []
        for e in kwargs.results:
            pdict[e.Serial].firmware = e.Firmware
            if 'sys/' in e.Dn: pdict[e.Serial].server_dn = (e.Dn).replace('sys/', '')
            elif e.ServerId == 0: pdict[e.Serial].server_dn = f'chassis-{e.ChassisId}-slot-{e.SlotId}'
            else: pdict[e.Serial].server_dn = f'rack-unit-{e.ServerId}'
            pdict[e.Serial].moid = e.Moid
            if e.get('ServiceProfile'): pdict[e.Serial].server_profile = e.ServiceProfile
            if not e.ManagementMode == 'IntersightStandalone':
                pdict[e.Serial].registered_moid = e.RegisteredDevice.Moid
                registered_devices.append(e.RegisteredDevice.Moid)
        if len(registered_devices) > 0:
            domain_map = DotMap()
            parents    = []
            kwargs = kwargs | DotMap(method = 'get', names = list(numpy.unique(numpy.array(registered_devices))), uri = 'asset/DeviceRegistrations')
            kwargs = isight.api('registered_device').calls(kwargs)
            for e in kwargs.results:
                if e.get('ParentConnection'):
                    domain_map[e.Moid].hostname = None
                    domain_map[e.Moid].parent = e.ParentConnection.Moid
                    parents.append(e.ParentConnection.Moid)
                else:
                    domain_map[e.Moid].hostname = e.DeviceHostname
                    domain_map[e.Moid].parent = None
            if len(parents) > 0:
                kwargs.names = list(numpy.unique(numpy.array(parents)))
                kwargs.uri   = 'asset/DeviceRegistrations'
                kwargs       = isight.api('registered_device').calls(kwargs)
                for e in kwargs.results:
                    for k, v in domain_map.items():
                        if v.get('parent'):
                            if v.parent == e.Moid: domain_map[k].hostname = e.DeviceHostname
            for k, v in pdict.items():
                if v.get('registered_moid'): pdict[k].domain = domain_map[v.registered_moid].hostname[0]
        hardware_moids = DotMap()
        for k,v in pdict.items(): hardware_moids[v.moid].serial = k
        kwargs.names      = list(hardware_moids.keys())
        kwargs.uri        = 'cond/HclStatuses'
        kwargs            = isight.api('hcl_status').calls(kwargs)
        hcl_results       = kwargs.results
        names             = "', '".join(kwargs.names).strip("', '")
        kwargs.api_filter = f"AssociatedServer.Moid in ('{names}')"
        kwargs.uri        = kwargs.ezdata['profiles.server'].intersight_uri
        kwargs            = isight.api('hcl_status').calls(kwargs)
        profile_results   = kwargs.results
        for k,v in pdict.items():
            indx = next((index for (index, d) in enumerate(hcl_results) if d['ManagedObject']['Moid'] == v.moid), None)
            if not indx == None:
                pdict[k].hcl_check  = hcl_results[indx].SoftwareStatus
                pdict[k].hcl_status = hcl_results[indx].Status
                pdict[k].hcl_reason = hcl_results[indx].ServerReason
            indx = next((index for (index, d) in enumerate(profile_results) if d['AssociatedServer']['Moid'] == v.moid), None)
            if not indx == None: pdict[k].server_profile = profile_results[indx].Name

        if len(pdict) > 0:
            kwargs.timezone = questions.prompt_user.for_timezone(kwargs)
            kwargs = tools(self.type).setup_local_time(kwargs)
            # Build Named Style Sheets for Workbook
            kwargs = tools(self.type).workbook_styles(kwargs)
            workbook = f'HCL-Inventory-{kwargs.time_short}.xlsx'
            wb = kwargs.wb
            ws = wb.active
            ws.title = 'Inventory List'
            column_headers = [
                'Domain','Model','Serial','Server','Profile','Firmware','vCenter','Cluster','Hostname', 'ESX Version', 'ESX Build',
                'HCL Component Status', 'HCL Status', 'HCL Reason', 'UCS Tools Name', 'UCS Tools Install Date', 'UCS Tools Version']
            for i in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = 30
            cLength = len(column_headers)
            ws_header = f'Collected UCS Data on {kwargs.time_long}'
            data = [ws_header]
            ws.append(data)
            ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
            for cell in ws['1:1']: cell.style = 'heading_1'
            ws.append(column_headers)
            for cell in ws['2:2']: cell.style = 'heading_2'
            ws_row_count = 3
            # Populate the Columns with Server Inventory
            for key, value in pdict.items():
                # Add the Columns to the Spreadsheet
                data = []
                for k,v in value.items():
                    if not re.search('moid|registered_moid', k): data.append(v)
                ws.append(data)
                for cell in ws[ws_row_count:ws_row_count]:
                    if ws_row_count % 2 == 0: cell.style = 'odd'
                    else: cell.style = 'even'
                ws_row_count += 1
            # Save the Workbook
            wb.save(filename=workbook)

    #=========================================================================
    # Function - Intersight Inventory
    #=========================================================================
    def inventory(self, kwargs):
        kwargs.org = 'default'
        # kwargs.firmware = DotMap(json.load(open(f'{kwargs.script_path}/QA/firmware.json', 'r')))
        # kwargs.domains  = DotMap(json.load(open(f'{kwargs.script_path}/QA/domains.json', 'r')))
        # kwargs.chassis  = DotMap(json.load(open(f'{kwargs.script_path}/QA/chassis.json', 'r')))
        # kwargs.servers  = DotMap(json.load(open(f'{kwargs.script_path}/QA/servers.json', 'r')))
        #=====================================================================
        # Running Firmware
        #=====================================================================
        kwargs.api_filter = 'ignore'
        kwargs = isight.api('firmware').running_firmware(kwargs)
        with open('firmware.json', 'w') as json_file:
            json.dump(kwargs.firmware, json_file, indent=4)
        #=====================================================================
        # Domain Inventory
        #=====================================================================
        kwargs.api_filter = f"PlatformType in ('UCSFI', 'UCSFIISM')"
        kwargs = isight.api('domains').domain_device_registrations(kwargs)
        kwargs.api_filter = f"SwitchType eq 'FabricInterconnect'"
        kwargs = isight.api('domains').domain_network_elements(kwargs)
        for e in ['switch_profiles', 'cluster_profiles']:
            kwargs.api_filter = 'ignore'
            kwargs = eval(f"isight.api('domains').domain_{e}(kwargs)")
        # with open('domains.json', 'w') as json_file:
        #     json.dump(kwargs.domains, json_file, indent=4)
        #=====================================================================
        # Chassis Inventory
        #=====================================================================
        for e in ['equipment', 'io_cards', 'profiles']:
            kwargs.api_filter = 'ignore'
            kwargs = eval(f"isight.api('chassis').chassis_{e}(kwargs)")
        # with open('chassis.json', 'w') as json_file:
        #     json.dump(kwargs.chassis, json_file, indent=4)
        #=====================================================================
        # Server Inventory
        #=====================================================================
        for e in ['compute', 'children_equipment', 'profiles', 'virtual_drives']:
            kwargs.api_filter = 'ignore'
            kwargs = eval(f"isight.api('server').server_{e}(kwargs)")
        # with open('servers.json', 'w') as json_file:
        #     json.dump(kwargs.servers, json_file, indent=4)
        #=====================================================================
        # Contract Inventory
        #=====================================================================
        kwargs.api_filter = 'ignore'
        kwargs = isight.api('contracts').inventory_contracts(kwargs)
        #=====================================================================
        # Add Server/PCI Node Data to Chassis Dictionary
        #=====================================================================
        for k,v in kwargs.servers.items():
            vkeys = list(v.keys())
            if 'pci_node' in vkeys and v.pci_node != None:
                pdict = DotMap(contract = v.contract, model = v.pci_node.model, serial = v.pci_node.serial)
                kwargs.chassis[v.chassis].slot[str(v.pci_node.slot)] = pdict
            if 'chassis' in vkeys:
                sdict = DotMap(contract = v.contract, model = v.model, serial = v.serial)
                kwargs.chassis[v.chassis].slot[str(v.slot)] = sdict
                if re.search('410|480', v.model): kwargs.chassis[v.chassis].slot[str(v.slot+1)] = sdict
        #=====================================================================
        # Sort the Dictionaries
        #=====================================================================
        kwargs.chassis = DotMap(sorted(kwargs.chassis.items(), key=lambda ele: ele[1].chassis_name))
        kwargs.domains = DotMap(sorted(kwargs.domains.items(), key=lambda ele: ele[1].device_hostname))
        kwargs.servers = DotMap(sorted(kwargs.servers.items(), key=lambda ele: ele[1].server_name))
        inventory      = DotMap(chassis = kwargs.chassis, domains = kwargs.domains, servers = kwargs.servers)
        with open('inventory.json', 'w') as json_file:
            json.dump(inventory, json_file, indent=4)

        #if len(kwargs.servers) > 0: kwargs = tools('workbook').inventory_workbook(kwargs)
        return kwargs

    #=========================================================================
    # Function - Inventory Workbook
    #=========================================================================
    def inventory_workbook(self, kwargs):
        #=====================================================================
        # Workbook Setup
        #=====================================================================
        #kwargs.timezone = questions.prompt_user.for_timezone(kwargs)
        kwargs.timezone = 'America/New_York'
        kwargs          = tools(self.type).setup_local_time(kwargs)
        workbook        = f'UCS-Inventory-Collector-{kwargs.time_short}.xlsx'
        kwargs          = tools(self.type).workbook_styles(kwargs)
        kwargs          = tools('domain').inventory_worksheet_domain(kwargs)
        #=====================================================================
        # Setup Domain Inventory WorkSheet
        #=====================================================================
        column_headers = ['Organization', 'Name', 'Model', 'Serial A', 'Serial B', 'Management Mode', 'Profile', 'Firmware']
        column_width   = DotMap()
        ws             = kwargs.wb.active
        ws.title       = 'Domain Inventory'
        for e in column_headers: column_width[e] = len(e)
        for k in list(kwargs.domains.keys()):
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'serial_a': elength = len(kwargs.domains[k].serial[0])
                elif key == 'serial_b': elength = len(kwargs.domains[k].serial[1])
                elif key == 'firmware': elength = len(kwargs.domains[k].firmware[0])
                else:  elength = len(kwargs.domains[k][key])
                if column_width[e] < elength: column_width[e] = elength
        for i in range(0,len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = column_width[column_headers[i]]
        ws_header     = f'UCS Domain Inventory {kwargs.time_long}'
        data          = [ws_header]; ws.append(data)
        column_length = len(column_headers)
        ws.merge_cells(f'A1:{chr(ord("@")+column_length)}1')
        for cell in ws['1:1']: cell.style = 'heading_1'
        ws.append(column_headers)
        for cell in ws['2:2']: cell.style = 'heading_2'
        ws_row_count = 3
        #=================================================================
        # Populate the Worksheet with Domain Inventory
        #=================================================================
        for k in list(kwargs.domains.keys()):
            data = []
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'serial_a': data.append(kwargs.domains[k].serial[0])
                elif key == 'serial_b': data.append(kwargs.domains[k].serial[1])
                elif key == 'firmware': data.append(kwargs.domains[k].firmware[0])
                else: data.append(kwargs.domains[k][key])
            ws.append(data)
            for cell in ws[ws_row_count:ws_row_count]:
                if ws_row_count % 2 == 0: cell.style = 'odd'
                else: cell.style = 'even'
            ws_row_count += 1
        #=====================================================================
        # Save the Workbook and return kwargs
        #=====================================================================
        kwargs.wb.save(filename=workbook)
        return kwargs

    #=========================================================================
    # Function - Inventory Workbook - Chassis Worksheet
    #=========================================================================
    def inventory_worksheet_chassis(self, kwargs):
        column_headers = ['Organization', 'Name', 'Model', 'Serial', 'Management Mode', 'Profile', 'IFM Model', 'IFM A Serial', 'IFM A Version',
                          'IFM B Serial', 'IFM B Version', 'X Fabric Model', 'X Fabric A Serial', 'X Fabric B Serial']
        for x in range(1,9): column_headers.append(f'Slot {x}')
        column_width   = DotMap()
        ws             = kwargs.wb.active
        ws.title       = 'Chassis Inventory'
        for e in column_headers: column_width[e] = len(e)
        for k in list(kwargs.chassis.keys()):
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'ifm_model': elength = len(kwargs.chassis[k].if_modules[0].model)
                elif re.search('(IFM|X Fabric) . Serial'): elength = len(kwargs.chassis[k].if_modules[0].version)
                elif key == 'firmware': elength = len(kwargs.chassis[k].firmware[0])
                else:  elength = len(kwargs.chassis[k][key])
                if column_width[e] < elength: column_width[e] = elength
        for i in range(0,len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = column_width[column_headers[i]] + 5
        ws_header     = f'UCS Domain Inventory {kwargs.time_long}'
        data          = [ws_header]; ws.append(data)
        column_length = len(column_headers)
        ws.merge_cells(f'A1:{chr(ord("@")+column_length)}1')
        for cell in ws['1:1']: cell.style = 'heading_1'
        ws.append(column_headers)
        for cell in ws['2:2']: cell.style = 'heading_2'
        ws_row_count = 3
        #=====================================================================
        # Populate the Worksheet with Domain Inventory
        #=====================================================================
        for k in list(kwargs.domains.keys()):
            data = []
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'serial_a': data.append(kwargs.domains[k].serial[0])
                elif key == 'serial_b': data.append(kwargs.domains[k].serial[1])
                elif key == 'firmware': data.append(kwargs.domains[k].firmware[0])
                else: data.append(kwargs.domains[k][key])
            ws.append(data)
            for cell in ws[ws_row_count:ws_row_count]:
                if ws_row_count % 2 == 0: cell.style = 'odd'
                else: cell.style = 'even'
            ws_row_count += 1
        #=====================================================================
        # Return kwargs
        #=====================================================================
        return kwargs

    #=========================================================================
    # Function - Inventory Workbook - Domain Worksheet
    #=========================================================================
    def inventory_worksheet_domain(self, kwargs):
        column_headers = ['Organization', 'Name', 'Model', 'Serial', 'Management Mode', 'Profile', 'Firmware']
        column_width   = DotMap()
        ws             = kwargs.wb.active
        ws.title       = 'Domain Inventory'
        for e in column_headers: column_width[e] = len(e)
        for k in list(kwargs.domains.keys()):
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'serial_a': elength = len(kwargs.domains[k].serial[0])
                elif key == 'serial_b': elength = len(kwargs.domains[k].serial[1])
                elif key == 'firmware': elength = len(kwargs.domains[k].firmware[0])
                else:  elength = len(kwargs.domains[k][key])
                if column_width[e] < elength: column_width[e] = elength
        for i in range(0,len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = column_width[column_headers[i]] + 5
        ws_header     = f'UCS Domain Inventory {kwargs.time_long}'
        data          = [ws_header]; ws.append(data)
        column_length = len(column_headers)
        ws.merge_cells(f'A1:{chr(ord("@")+column_length)}1')
        for cell in ws['1:1']: cell.style = 'heading_1'
        ws.append(column_headers)
        for cell in ws['2:2']: cell.style = 'heading_2'
        ws_row_count = 3
        #=====================================================================
        # Populate the Worksheet with Domain Inventory
        #=====================================================================
        for k in list(kwargs.domains.keys()):
            data = []
            for e in column_headers:
                key = snakecase(e).replace('__', '_')
                if   key == 'serial_a': data.append(kwargs.domains[k].serial[0])
                elif key == 'serial_b': data.append(kwargs.domains[k].serial[1])
                elif key == 'firmware': data.append(kwargs.domains[k].firmware[0])
                else: data.append(kwargs.domains[k][key])
            ws.append(data)
            for cell in ws[ws_row_count:ws_row_count]:
                if ws_row_count % 2 == 0: cell.style = 'odd'
                else: cell.style = 'even'
            ws_row_count += 1
        #=====================================================================
        # Return kwargs
        #=====================================================================
        return kwargs

    #=========================================================================
    # Function - Profiles Update
    #=========================================================================
    def profiles_deploy(self, kwargs):
        #=====================================================================
        # Get Physical Devices
        #=====================================================================
        kwargs.names = []
        for e in kwargs.profile_results:
            if   e.get('AssignedServer'):  kwargs.names.append(e.AssignedServer.Moid)
            elif e.get('AssignedChassis'): kwargs.names.append(e.AssignedChassis.Moid)
            elif e.get('AssignedSwitch'):  kwargs.names.append(e.AssignedSwitch.Moid)
        kwargs = kwargs | DotMap(method = 'get', uri = kwargs.ezdata[self.type].intersight_uri_serial)
        kwargs = isight.api('moid_filter').calls(kwargs)
        phys_devices = kwargs.results
        profiles     = []
        #=====================================================================
        # Build Dictionary and Deploy Profiles
        #=====================================================================
        if re.search('chassis|server', self.type):
            for e in kwargs.profile_names:
                indx = next((index for (index, d) in enumerate(kwargs.profile_results) if d['Name'] == e), None)
                kwargs.isight[kwargs.org].profiles[self.type][e] = kwargs.profile_results[indx].Moid
                if self.type == 'server': sname = 'AssignedServer'
                if self.type == 'chassis': sname = 'AssignedChassis'
                sindx = next((index for (index, d) in enumerate(phys_devices) if d['Moid'] == kwargs.profile_results[indx][sname].Moid), None)
                profiles.append(DotMap(action='Deploy',name=e,serial_number=phys_devices[sindx].Serial))
            if len(profiles) > 0: kwargs = isight.imm(self.type).profile_chassis_server_deploy(profiles, kwargs)
        else:
            for dp in kwargs.domain_names:
                indx = next((index for (index, d) in enumerate(kwargs.domain_results) if d['Name'] == dp), None)
                kwargs.isight[kwargs.org].profiles[self.type][dp] = kwargs.domain_results[indx].Moid
                switch_profiles = [e for e in kwargs.profile_results if e.SwitchClusterProfile.Moid == kwargs.domain_results[indx].Moid]
                serials         = []
                for e in switch_profiles:
                    kwargs.isight[kwargs.org].profiles['switch'][e.Name] = e.Moid
                    sname = 'AssignedSwitch'
                    sindx = next((index for (index, d) in enumerate(phys_devices) if d['Moid'] == e[sname].Moid), None)
                    serials.append(phys_devices[sindx])
                serials = sorted(serials, key=lambda ele: ele.SwitchId)
                profiles.append(DotMap(action='Deploy',name=dp,serial_numbers=[e.Serial for e in serials]))
            if len(profiles) > 0: kwargs = isight.imm(self.type).profile_domain_deploy(profiles, kwargs)
        #=====================================================================
        # Return kwargs
        #=====================================================================
        return kwargs

    #=========================================================================
    # Function - Profiles Update
    #=========================================================================
    def profiles_update(self, kwargs):
        #=====================================================================
        # Obtain Profile Data.
        #=====================================================================
        org_moid = kwargs.org_moids[kwargs.org].moid
        if kwargs.get('server_type') and type(kwargs.server_type) == str: utype = kwargs.server_type
        else: utype = self.type
        if re.search('^server(_template)?$', utype):
            kwargs.api_filter   = f"Organization.Moid eq '{org_moid}' and TargetPlatform eq '{kwargs.target_platform}'"
        else: kwargs.api_filter = f"Organization.Moid eq '{org_moid}'"
        kwargs = kwargs | DotMap(method = 'get', uri = kwargs.ezdata[utype].intersight_uri)
        kwargs = isight.api(utype).calls(kwargs)
        profiles = kwargs.pmoids
        if len(profiles.toDict()) == 0: isight.empty_results(kwargs)
        kwargs.profile_results = kwargs.results
        #=====================================================================
        # Request from User Which Profiles to Apply this to.
        #=====================================================================
        kwargs.profile_names = sorted(list(profiles.keys()))
        kwargs.jdata = DotMap(
            default      = kwargs.profile_names[0],
            description  = f'Select the `{utype}` Profiles to update.',
            enum         = kwargs.profile_names,
            multi_select = True,
            title        = f'{self.type} Profiles')
        kwargs.profile_names = ezfunctions.variable_prompt(kwargs)
        if kwargs.target_platform == 'domain':
            kwargs.domain_names    = kwargs.profile_names
            kwargs.domain_results  = kwargs.profile_results
            kwargs.names           = [profiles[e].moid for e in kwargs.profile_names]
            kwargs.uri             = kwargs.ezdata[utype].intersight_uri_switch
            kwargs                 = isight.api('switch_profiles').calls(kwargs)
            profiles               = kwargs.pmoids
            kwargs.profile_names   = list(profiles.keys())
            kwargs.profile_results = kwargs.results
        #=====================================================================
        # Attach the Policies to the Profiles.
        #=====================================================================
        kwargs.bulk_list = []
        for i in kwargs.profile_names:
            original_bucket = profiles[i].policy_bucket
            new_bucket = []
            for e in original_bucket:
                e.pop('ClassId'); e.pop('link')
                indx = next((index for (index, d) in enumerate(kwargs.policy_bucket) if d['ObjectType'] == e['ObjectType']), None)
                if indx == None: new_bucket.append(e.toDict())
                else: new_bucket.append(kwargs.policy_bucket[indx].toDict())
            policy_bucket = sorted(new_bucket, key=lambda ele: ele['ObjectType'])
            api_body = {"Name":i,"PolicyBucket":policy_bucket,'pmoid':profiles[i].moid}
            if len(kwargs.pools) > 0:
                for e in kwargs.pools:
                    if e.ObjectType == 'resourcepool.Pool': api_body.update({'ServerAssignmentMode':'Pool','ServerPool':e.toDict()})
                    elif e.ObjectType == 'uuidpool.Pool':   api_body.update({'UuidAddressType':'Pool','UuidPool':e.toDict()})
            kwargs.bulk_list.append(api_body)
        #=====================================================================
        # Send the Bulk Request
        #=====================================================================
        if len(kwargs.bulk_list) > 0:
            kwargs.uri = kwargs.ezdata[utype].intersight_uri
            kwargs     = isight.imm(utype).bulk_request(kwargs)
        #=====================================================================
        # Prompt User to Deploy Profile(s)
        #=====================================================================
        if len(kwargs.bulk_list) > 0 and re.search('^chassis|domain|server$', utype):
            kwargs.jdata             = kwargs.ezwizard.setup.properties.deploy_profiles
            kwargs.jdata.description = kwargs.jdata.description.replace('server', utype)
            deploy_profiles          = ezfunctions.variable_prompt(kwargs)
            if deploy_profiles == True: tools(utype).profiles_deploy(kwargs)
        return kwargs

    #=========================================================================
    # Function - Setup Local Time
    #=========================================================================
    def setup_local_time(self, kwargs):
        kwargs.datetime   = datetime.now(pytz.timezone(kwargs.timezone))
        kwargs.time_short = kwargs.datetime.strftime('%Y-%m-%d-%H-%M')
        kwargs.time_long  = kwargs.datetime.strftime('%Y-%m-%d %H:%M:%S %Z %z')
        return kwargs

    #=========================================================================
    # Function - Server Inventory
    #=========================================================================
    def server_identities(self, kwargs):
        #=====================================================================
        # Get Device Registrations and build domains/servers
        #=====================================================================
        kwargs.org = 'default'
        kwargs.api_filter = f"PlatformType in ('IMCBlade', 'IMCRack', 'IMCM4', 'IMCM5', 'IMCM6', 'IMCM7', 'IMCM8', 'IMCM9', 'UCSFI', 'UCSFIISM')"
        kwargs = kwargs | DotMap(method = 'get', uri = 'asset/DeviceRegistrations')
        kwargs = isight.api('device_registration').calls(kwargs)
        for e in kwargs.results:
            if re.search('UCSFI(ISM)?', e.PlatformType):
                kwargs.domains[e.Moid] = DotMap(name = e.DeviceHostname[0], serials = e.Serial, servers = DotMap(), type = e.PlatformType)
            elif re.search('IMC', e.PlatformType):
                parent = 'none'
                if e.get('ParentConnection'): parent = e.ParentConnection.Moid
                kwargs.servers[e.Serial[0]] = DotMap(dict(parent = parent, server_name = e.DeviceHostname[0]))
        #=====================================================================
        # Get Physical Summaries and extend servers dict
        #=====================================================================
        kwargs.api_filter = 'ignore'
        kwargs.uri        = 'compute/PhysicalSummaries'
        kwargs            = isight.api('physical_summaries').calls(kwargs)
        for e in kwargs.results:
            kwargs.servers[e.Serial] = DotMap(dict(kwargs.servers[e.Serial].toDict(), **dict(
                chassis_id      = e.ChassisId,
                domain          = 'Standalone',
                hardware_moid   = e.Moid,
                mgmt_ip_address = e.MgmtIpAddress,
                mgmt_mode       = e.ManagementMode,
                model           = e.Model,
                moid            = '',
                name            = e.Name,
                object_type     = e.SourceObjectType,
                organization    = 'none',
                platform_type   = e.PlatformType,
                power_state     = e.OperPowerState,
                registration    = e.RegisteredDevice.Moid,
                server_dn       = e.Dn,
                server_id       = e.ServerId,
                server_profile  = 'unassigned',
                slot            = e.SlotId,
                wwnn            = 'unassigned')))
            if len(e.ServiceProfile) > 0: kwargs.servers[e.Serial].server_profile = e.ServiceProfile
        for k in list(kwargs.servers.keys()):
            if not kwargs.servers[k].get('hardware_moid'): kwargs.servers.pop(k)
            else: kwargs.hardware_moids[kwargs.servers[k].hardware_moid] = k
        #=====================================================================
        # Get Intersight Server Profiles
        #=====================================================================
        kwargs.api_filter = 'ignore'
        kwargs.uri        = 'server/Profiles'
        kwargs            = isight.api('server').calls(kwargs)
        if kwargs.results == None: prRed('empty results.  Exiting script...'); sys.exit(1)
        profile_moids = []
        for e in kwargs.results:
            if e.AssociatedServer != None:
                serial = kwargs.hardware_moids[e.AssignedServer.Moid]
                kwargs.servers[serial].moid           = e.Moid
                kwargs.servers[serial].organization   = dict(name = kwargs.org_names[e.Organization.Moid], moid = e.Organization.Moid)
                kwargs.servers[serial].server_profile = e.Name
                profile_moids.append(e.Moid)
        if len(profile_moids) > 0:
            #=================================================================
            # Get WWPN Leases, vHBAs, vNICs, and QoS Policies
            #=================================================================
            kwargs.api_filter = "PoolPurpose eq 'WWNN'"
            kwargs.uri   = 'fcpool/Leases'
            kwargs       = isight.api('wwnn_pool_leases').calls(kwargs)
            wwnn_leases  = kwargs.results
            kwargs.api_filter = 'ignore'
            kwargs.uri   = kwargs.ezdata['san_connectivity.vhbas'].intersight_uri
            kwargs       = isight.api('profile_moid').calls(kwargs)
            vhbas        = kwargs.results
            kwargs.api_filter = 'ignore'
            kwargs.uri   = kwargs.ezdata['lan_connectivity.vnics'].intersight_uri
            kwargs       = isight.api('profile_moid').calls(kwargs)
            vnics        = kwargs.results
            kwargs.names = list(numpy.unique(numpy.array([e.EthQosPolicy.Moid for e in vnics])))
            kwargs.uri   = kwargs.ezdata.ethernet_qos.intersight_uri
            kwargs       = isight.api('moid_filter').calls(kwargs)
            qos_policies = DotMap()
            for e in kwargs.results: qos_policies[e.Moid] = DotMap(mtu = e.Mtu, name = e.Name)
            #=================================================================
            # Add vHBAs/vNICs to servers dict
            #=================================================================
            profile_wwnn_leases = []
            for e in wwnn_leases:
                if not e.AssignedToEntity == None: profile_wwnn_leases.append(e)
            profile_vhbas = DotMap()
            profile_vnics = DotMap()
            for e in vhbas:
                if not e.Profile == None:
                    if not type(profile_vhbas[e.Profile.Moid]) == list: profile_vhbas[e.Profile.Moid] = []
                    profile_vhbas[e.Profile.Moid].append(e)
            for e in vnics:
                if not e.Profile == None:
                    if not type(profile_vnics[e.Profile.Moid]) == list: profile_vnics[e.Profile.Moid] = []
                    profile_vnics[e.Profile.Moid].append(e)
            vhba_keys = list(profile_vhbas.keys())
            vnic_keys = list(profile_vnics.keys())
            for k in list(kwargs.servers.keys()):
                kwargs.servers[k].vhbas = []
                kwargs.servers[k].vnics = []
                indx = next((index for (index, d) in enumerate(profile_wwnn_leases) if d.AssignedToEntity.Moid == kwargs.servers[k].moid), None)
                if indx != None: kwargs.servers[k].wwnn = wwnn_leases[indx].WwnId
                if kwargs.servers[k].moid in vhba_keys:
                    for e in profile_vhbas[kwargs.servers[k].moid]:
                        if e.WwpnAddressType == 'STATIC':
                            kwargs.servers[k].vhbas.append({'name': e.Name, 'order': e.Order, 'switch_id':e.Placement.SwitchId, 'vif_id': e.VifId, 'wwpn_address':e.StaticWwpnAddress})
                        else: kwargs.servers[k].vhbas.append({'name': e.Name, 'order': e.Order, 'switch_id':e.Placement.SwitchId, 'vif_id': e.VifId,'wwpn_address':e.Wwpn})
                if kwargs.servers[k].moid in vnic_keys:
                    for e in profile_vnics[kwargs.servers[k].moid]:
                        if e.MacAddressType == 'STATIC':
                            kwargs.servers[k].vnics.append({'name': e.Name, 'mac_address':e.StaticMacAddress,'mtu':qos_policies[e.EthQosPolicy.Moid].mtu, 'order': e.Order, 'vif_id': e.VifId})
                        else: kwargs.servers[k].vnics.append({'name': e.Name, 'mac_address':e.MacAddress,'mtu':qos_policies[e.EthQosPolicy.Moid].mtu, 'order': e.Order, 'vif_id': e.VifId})
        #=====================================================================
        # BUILD Workbooks
        #=====================================================================
        if len(kwargs.servers) > 0:
            #=================================================================
            # Sort server dictionary
            #=================================================================
            kwargs.servers = DotMap(dict(sorted(kwargs.servers.items(), key=lambda ele: ele[1].server_profile)))
            for k, v in kwargs.servers.items():
                kwargs.servers[k].vnics = sorted(v.vnics, key=lambda ele: ele.order)
                kwargs.servers[k].vhbas = sorted(v.vhbas, key=lambda ele: ele.order)
                if v.platform_type == 'UCSFI': kwargs.servers[k].domain = kwargs.domains[v.registration].name
                elif v.mgmt_mode != 'IntersightStandalone': kwargs.servers[k].domain = kwargs.domains[v.parent].name
            kwargs.timezone = questions.prompt_user.for_timezone(kwargs)
            kwargs          = tools(self.type).setup_local_time(kwargs)
            #=================================================================
            # Build Workbooks and Style Sheets
            #=================================================================
            workbook   = f'UCS-Server-Identities-{kwargs.time_short}.xlsx'
            kwargs = tools(self.type).workbook_styles(kwargs)
            wb = kwargs.wb; ws = wb.active; ws.title = 'Identity List'
            #=================================================================
            # Full Inventory Workbook
            #=================================================================
            if kwargs.args.full_identities:
                #=============================================================
                # Create Column Headers - Extend with server dict
                #=============================================================
                column_headers = ['Domain','Profile','Server','Serial','VIF','WWNN']
                vhba_list = []; vnic_list = []
                for i in list(kwargs.servers.keys()):
                    for e in kwargs.servers[i].vhbas:
                        if not e.name in vhba_list: vhba_list.append(e.name)
                    for e in kwargs.servers[i].vnics:
                        if not e.name in vnic_list: vnic_list.append(e.name)
                column_headers = column_headers + vhba_list + vnic_list
                for i in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = 30
                cLength = len(column_headers)
                ws_header = f'Collected UCS Data on {kwargs.time_long}'
                data = [ws_header]
                ws.append(data)
                ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
                for cell in ws['1:1']: cell.style = 'heading_1'
                ws.append(column_headers)
                for cell in ws['2:2']: cell.style = 'heading_2'
                ws_row_count = 3
                #=============================================================
                # Populate the Worksheet with Server Inventory
                #=============================================================
                for k, v in kwargs.servers.items():
                    data = []
                    for i in column_headers:
                        column_count = 0
                        if   i == 'Domain':  data.append(v.domain); column_count += 1
                        elif i == 'Profile': data.append(v.server_profile); column_count += 1
                        elif i == 'Serial':  data.append(k); column_count += 1
                        elif i == 'WWNN':    data.append(v.wwnn); column_count += 1
                        elif i == 'Server':
                            if not 'sys' in v.server_dn:
                                if len(v.chassis_id) > 0: server_dn = f'sys/chassis-{v.chassis_id}/blade-{v.slot}'
                                else: server_dn = f'sys/rack-unit-{v.server_id}'
                            else: server_dn = v.server_dn
                            if len(server_dn) == 0: data.append(''); column_count += 1
                            else: data.append(server_dn); column_count += 1
                        else:
                            for e in v.vhbas:
                                if i == e.name:
                                    data.append(e.wwpn_address); column_count += 1
                                    data[4] = e.vif_id  # Update VIF column
                            for e in v.vnics:
                                if i == e.name:
                                    data.append(e.mac_address); column_count += 1
                                    data[4] = e.vif_id  # Update VIF column
                        if column_count == 0:
                            if i == 'VIF': data.append('unassigned')
                            else: data.append('Not Configured')
                    ws.append(data)
                    for cell in ws[ws_row_count:ws_row_count]:
                        if ws_row_count % 2 == 0: cell.style = 'odd'
                        else: cell.style = 'even'
                    ws_row_count += 1
            #=================================================================
            # WWPN Inventory Workbook
            #=================================================================
            else:
                #=============================================================
                # Create Column Headers - Extend with server dict
                #=============================================================
                column_headers = ['Profile','Serial']
                vhba_list = []
                for k,v in kwargs.servers.items():
                    if v.wwnn != 'unassigned':
                        if not 'WWNN' in column_headers: column_headers.append('WWNN')
                    if v.get('vhbas'):
                        for e in v.vhbas:
                            if not e.name in vhba_list: vhba_list.append(e.name)
                column_headers= column_headers + vhba_list
                for x in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+x+1)].width = 30
                cLength = len(column_headers)
                ws_header = f'Collected UCS Data on {kwargs.time_long}'
                data = [ws_header]
                ws.append(data)
                ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
                for cell in ws['1:1']: cell.style = 'heading_1'
                ws.append(column_headers)
                for cell in ws['2:2']: cell.style = 'heading_2'
                ws_row_count = 3
                #=============================================================
                # Populate the Worksheet with Server Inventory
                #=============================================================
                for k, v in kwargs.servers.items():
                    data = []
                    for i in column_headers:
                        column_count = 0
                        if   i == 'Profile': data.append(v.server_profile); column_count += 1
                        elif i == 'Serial':  data.append(k); column_count += 1
                        elif i == 'WWNN':    data.append(v.wwnn); column_count += 1
                        else:
                            if v.get('vhbas'):
                                for e in v.vhbas:
                                    if i == e.name: data.append(e.wwpn_address); column_count += 1
                        if column_count == 0: data.append('Not Configured')
                    ws.append(data)
                    for cell in ws[ws_row_count:ws_row_count]:
                        if ws_row_count % 2 == 0: cell.style = 'odd'
                        else: cell.style = 'even'
                    ws_row_count += 1
            #=================================================================
            # Save the Workbook
            #=================================================================
            wb.save(filename=workbook)
            pcolor.LightPurple(f'\n{"-"*108}\n')
            pcolor.Yellow(f'  Saved Server Inventory to `{os.getcwd()}/{workbook}`')
            pcolor.LightPurple(f'\n{"-"*108}')

    #=========================================================================
    # Function - Build Named Style Sheets for Workbook
    #=========================================================================
    def workbook_styles(self, kwargs):
        wb = openpyxl.Workbook()
        bd1 = Side(style="thick", color="0070C0")
        bd2 = Side(style="medium", color="0070C0")
        heading_1 = NamedStyle(name="heading_1")
        heading_1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        heading_1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
        heading_1.fill = PatternFill("solid", fgColor="305496")
        heading_1.font = Font(bold=True, size=15, color="FFFFFF")
        heading_2 = NamedStyle(name="heading_2")
        heading_2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        heading_2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
        heading_2.font = Font(bold=True, size=15, color="44546A")
        even = NamedStyle(name="even")
        even.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        even.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
        even.font = Font(bold=False, size=12, color="44546A")
        odd = NamedStyle(name="odd")
        odd.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
        odd.fill = PatternFill("solid", fgColor="D9E1F2")
        odd.font = Font(bold=False, size=12, color="44546A")
        wb.add_named_style(heading_1)
        wb.add_named_style(heading_2)
        wb.add_named_style(even)
        wb.add_named_style(odd)
        kwargs.wb = wb
        return kwargs
