#!/usr/bin/env python3

from git import cmd, Repo
from openpyxl import load_workbook
from ordered_set import OrderedSet
import itertools
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import stdiomask
import validating
# from class_policies_domain import policies_domain
from textwrap import fill

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

#======================================================
# Function - Prompt User for the api_key
#======================================================
def api_key(args):
    if args.api_key_id == None:
        key_loop = False
        while key_loop == False:
            question = stdiomask.getpass(f'The Intersight API Key was not entered as a command line option.\n'\
                'Please enter the Version 2 Intersight API key to use: ')

            if len(question) == 74:
                args.api_key_id = question
                key_loop = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  The API key length should be 74 characters.  Please Re-Enter.')
                print(f'\n-------------------------------------------------------------------------------------------\n')

    return args.api_key_id

#======================================================
# Function - Prompt User for the api_secret
#======================================================
def api_secret(args):
    secret_loop = False
    while secret_loop == False:
        if '~' in args.api_key_file:
            secret_path = os.path.expanduser(args.api_key_file)
        else:
            secret_path = args.api_key_file
        if not os.path.isfile(secret_path):
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! api_key_file not found.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            args.api_key_file = input(f'Please Enter the Path to the File containing the Intersight API Secret: ')
        else:
            secret_file = open(secret_path, 'r')
            if 'RSA PRIVATE KEY' in secret_file.read():
                secret_loop = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! api_key_file does not seem to contain the Private Key.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
    return secret_path

#======================================================
# Function - Format Policy Description
#======================================================
def choose_policy(policy, **templateVars):

    if 'policies' in policy:
        policy_short = policy.replace('policies', 'policy')
    elif 'pools' in policy:
        policy_short = policy.replace('pools', 'pool')
    elif 'templates' in policy:
        policy_short = policy.replace('templates', 'template')
    x = policy_short.split('_')
    policy_description = []
    for y in x:
        y = y.capitalize()
        policy_description.append(y)
    policy_description = " ".join(policy_description)
    policy_description = policy_description.replace('Ip', 'IP')
    policy_description = policy_description.replace('Ntp', 'NTP')
    policy_description = policy_description.replace('Snmp', 'SNMP')
    policy_description = policy_description.replace('Wwnn', 'WWNN')
    policy_description = policy_description.replace('Wwpn', 'WWPN')

    if len(policy) > 0:
        templateVars["policy"] = policy_description
        policy_short = policies_list(templateVars["policies"], **templateVars)
    else:
        policy_short = ""
    return policy_short

#======================================================
# Function - Count the Number of Keys
#======================================================
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

#======================================================
# Function - Prompt User with question - default No
#======================================================
def exit_default_no(policy_type):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [N]: ')
        if exit_answer == '' or exit_answer == 'N':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        elif exit_answer == 'Y':
            policy_loop = False
            configure_loop = False
            valid_exit = True
        else:
            print(f'\n------------------------------------------------------\n')
            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
            print(f'\n------------------------------------------------------\n')
    return configure_loop, policy_loop

#======================================================
# Function - Prompt User with question - default Yes
#======================================================
def exit_default_yes(policy_type):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [Y]: ')
        if exit_answer == '' or exit_answer == 'Y':
            policy_loop = False
            configure_loop = False
            valid_exit = True
        elif exit_answer == 'N':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        else:
            print(f'\n------------------------------------------------------\n')
            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
            print(f'\n------------------------------------------------------\n')
    return configure_loop, policy_loop

#======================================================
# Function - Prompt User with question
#======================================================
def exit_loop_default_yes(loop_count, policy_type):
    valid_exit = False
    while valid_exit == False:
        if loop_count % 2 == 0:
            exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [Y]: ')
        else:
            exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [N]: ')
        if (loop_count % 2 == 0 and exit_answer == '') or exit_answer == 'Y':
            policy_loop = False
            configure_loop = False
            loop_count += 1
            valid_exit = True
        elif not loop_count % 2 == 0 and exit_answer == '':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        elif exit_answer == 'N':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        else:
            print(f'\n------------------------------------------------------\n')
            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
            print(f'\n------------------------------------------------------\n')
    return configure_loop, loop_count, policy_loop

#======================================================
# Function - find the Keys for each Section
#======================================================
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

#======================================================
# Function - Assign the Variables to the Keys
#======================================================
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#======================================================
# Function - ipmi_key Function
#======================================================
def ipmi_key_function(**templateVars):
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  The ipmi_key Must be in Hexidecimal Format [a-fA-F0-9] and no longer than 40 characters.')
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        password1 = stdiomask.getpass(prompt='Enter the ipmi_key: ')
        password2 = stdiomask.getpass(prompt='Please re-enter ipmi_key: ')
        if not password1 == '':
            if password1 == password2:
                TF_VAR = 'TF_VAR_ipmi_key_1'
                os.environ[TF_VAR] = '%s' % (password1)
                templateVars["ipmi_key"] = 1
                valid = validating.ipmi_key_check(password1)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! The Keys did not match.  Please Re-enter the IPMI Key.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Value.  Please Re-enter the IPMI Key.')
            print(f'\n-------------------------------------------------------------------------------------------\n')

    return templateVars["ipmi_key"]

#======================================================
# Function - Local User Policy
#======================================================
def local_users_function(jsonData, easy_jsonData, inner_loop_count, **templateVars):
    local_users = []
    valid_users = False
    while valid_users == False:
        templateVars["multi_select"] = False
        jsonVars = jsonData['components']['schemas']['iam.EndPointUser']['allOf'][1]['properties']

        templateVars["Description"] = jsonVars['Name']['description']
        templateVars["varDefault"] = 'admin'
        templateVars["varInput"] = 'What is the Local username?'
        templateVars["varName"] = 'Local User'
        templateVars["varRegex"] = jsonVars['Name']['pattern']
        templateVars["minLength"] = 1
        templateVars["maxLength"] = jsonVars['Name']['maxLength']
        username = varStringLoop(**templateVars)

        templateVars["multi_select"] = False
        jsonVars = easy_jsonData['policies']['iam.LocalUserPasswordPolicy']
        templateVars["var_description"] = jsonVars['role']['description']
        templateVars["jsonVars"] = sorted(jsonVars['role']['enum'])
        templateVars["defaultVar"] = jsonVars['role']['default']
        templateVars["varType"] = 'User Role'
        role = variablesFromAPI(**templateVars)

        if templateVars["enforce_strong_password"] == True:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print('Enforce Strong Password is enabled so the following rules must be followed:')
            print('  - The password must have a minimum of 8 and a maximum of 20 characters.')
            print("  - The password must not contain the User's Name.")
            print('  - The password must contain characters from three of the following four categories.')
            print('    * English uppercase characters (A through Z).')
            print('    * English lowercase characters (a through z).')
            print('    * Base 10 digits (0 through 9).')
            print('    * Non-alphabetic characters (! , @, #, $, %, ^, &, *, -, _, +, =)\n\n')
        valid = False
        while valid == False:
            password1 = stdiomask.getpass(f'What is the password for {username}? ')
            password2 = stdiomask.getpass(f'Please re-enter the password for {username}? ')
            if not password1 == '':
                if password1 == password2:
                    if templateVars["enforce_strong_password"] == True:
                        valid = validating.strong_password(f"{username}'s password", password1, 8, 20)

                    else:
                        valid = validating.string_length(f'{username} password', password1, 1, 127)

                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Error!! The Passwords did not match.  Please Re-enter the password for {username}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please Re-enter the password for {username}.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        TF_VAR = 'TF_VAR_local_user_password_%s' % (inner_loop_count)
        os.environ[TF_VAR] = '%s' % (password1)
        password1 = inner_loop_count

        user_attributes = {
            'enabled':True,
            'password':inner_loop_count,
            'role':role,
            'username':username
        }
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'   enabled  = True')
        print(f'   password = "Sensitive"')
        print(f'   role     = "{role}"')
        print(f'   username = "{username}"')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            question = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if question == 'Y' or question == '':
                local_users.append(user_attributes)
                valid_exit = False
                while valid_exit == False:
                    loop_exit = input(f'Would You like to Configure another Local User?  Enter "Y" or "N" [N]: ')
                    if loop_exit == 'Y':
                        inner_loop_count += 1
                        valid_confirm = True
                        valid_exit = True
                    elif loop_exit == 'N' or loop_exit == '':
                        user_loop = True
                        valid_confirm = True
                        valid_exit = True
                        valid_users = True
                    else:
                        print(f'\n------------------------------------------------------\n')
                        print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                        print(f'\n------------------------------------------------------\n')

            elif question == 'N':
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Starting Local User Configuration Over.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                valid_confirm = True
            else:
                print(f'\n------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                print(f'\n------------------------------------------------------\n')

    return local_users,user_loop

#======================================================
# Function - Merge Easy IMM Repository to Dest Folder
#======================================================
def merge_easy_imm_repository(args, easy_jsonData, org):
    baseRepo = args.dir

    # Setup Operating Environment
    opSystem = platform.system()
    tfe_dir = 'tfe_modules'
    if opSystem == 'Windows': path_sep = '\\'
    else: path_sep = '/'
    git_url = "https://github.com/terraform-cisco-modules/terraform-intersight-easy-imm"
    if not os.path.isdir(tfe_dir):
        os.mkdir(tfe_dir)
        Repo.clone_from(git_url, tfe_dir)
    if not os.path.isfile(os.path.join(tfe_dir, 'README.md')):
        Repo.clone_from(git_url, tfe_dir)
    else:
        g = cmd.Git(tfe_dir)
        g.pull()

    folder_list = [
        f'{baseRepo}{path_sep}{org}{path_sep}policies',
        f'{baseRepo}{path_sep}{org}{path_sep}pools',
        f'{baseRepo}{path_sep}{org}{path_sep}profiles',
        f'{baseRepo}{path_sep}{org}{path_sep}ucs_domain_profiles'
    ]

    removeList = [
        'data_sources.tf',
        'locals.tf',
        'main.tf',
        'output.tf',
        'outputs.tf',
        'provider.tf',
        'README.md',
        'variables.tf',
    ]
    # Now Loop over the folders and merge the module files
    module_folders = ['policies', 'pools', 'profiles', 'ucs_domain_profiles']
    for folder in folder_list:
        for mod in module_folders:
            fsplit = folder.split(path_sep)
            if fsplit[-1] == mod:
                src_dir = os.path.join(tfe_dir, 'modules', mod)
                copy_files = os.listdir(src_dir)
                for fname in copy_files:
                    if not os.path.isdir(os.path.join(src_dir, fname)):
                        shutil.copy2(os.path.join(src_dir, fname), folder)
                
                # Identify the files 
                files = easy_jsonData['wizard']['files'][mod]
                for xRemove in removeList:
                    if xRemove in files:
                        files.remove(xRemove)
                terraform_fmt(files, folder, path_sep)

#======================================================
# Function - Naming Rule
#======================================================
def naming_rule(name_prefix, name_suffix, org):
    if not name_prefix == '':
        name = '%s_%s' % (name_prefix, name_suffix)
    else:
        name = '%s_%s' % (org, name_suffix)
    return name

#======================================================
# Function - Naming Rule Fabric Policy
#======================================================
def naming_rule_fabric(loop_count, name_prefix, org):
    if loop_count % 2 == 0:
        if not name_prefix == '':
            name = '%s_A' % (name_prefix)
        elif not org == 'default':
            name = '%s_A' % (org)
        else:
            name = 'Fabric_A'
    else:
        if not name_prefix == '':
            name = '%s_B' % (name_prefix)
        elif not org == 'default':
            name = '%s_B' % (org)
        else:
            name = 'Fabric_B'
    return name

#======================================================
# Function - NTP
#======================================================
def ntp_alternate():
    valid = False
    while valid == False:
        alternate_true = input('Do you want to Configure an Alternate NTP Server?  Enter "Y" or "N" [Y]: ')
        if alternate_true == 'Y' or alternate_true == '':
            alternate_ntp = input('What is your Alternate NTP Server? [1.north-america.pool.ntp.org]: ')
            if alternate_ntp == '':
                alternate_ntp = '1.north-america.pool.ntp.org'
            if re.search(r'[a-zA-Z]+', alternate_ntp):
                valid = validating.dns_name('Alternate NTP Server', alternate_ntp)
            else:
                valid = validating.ip_address('Alternate NTP Server', alternate_ntp)
        elif alternate_true == 'N':
            alternate_ntp = ''
            valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return alternate_ntp

#======================================================
# Function - NTP
#======================================================
def ntp_primary():
    valid = False
    while valid == False:
        primary_ntp = input('What is your Primary NTP Server [0.north-america.pool.ntp.org]: ')
        if primary_ntp == "":
            primary_ntp = '0.north-america.pool.ntp.org'
        if re.search(r'[a-zA-Z]+', primary_ntp):
            valid = validating.dns_name('Primary NTP Server', primary_ntp)
        else:
            valid = validating.ip_address('Primary NTP Server', primary_ntp)
    return primary_ntp

def policies_list(policies_list, **templateVars):
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if templateVars.get('optional_message'):
            print(templateVars["optional_message"])
        print(f'  {templateVars["policy"]} Options:')
        for i, v in enumerate(policies_list):
            i += 1
            if i < 10:
                print(f'     {i}. {v}')
            else:
                print(f'    {i}. {v}')
        if templateVars["allow_opt_out"] == True:
            print(f'     99. Do not assign a(n) {templateVars["policy"]}.')
        print(f'     100. Create a New {templateVars["policy"]}.')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        policyOption = input(f'Select the Option Number for the {templateVars["policy"]} to Assign to {templateVars["name"]}: ')
        if re.search(r'^[0-9]{1,3}$', policyOption):
            for i, v in enumerate(policies_list):
                i += 1
                if int(policyOption) == i:
                    policy = v
                    valid = True
                    return policy
                elif int(policyOption) == 99:
                    policy = ''
                    valid = True
                    return policy
                elif int(policyOption) == 100:
                    policy = 'create_policy'
                    valid = True
                    return policy

            if int(policyOption) == 99:
                policy = ''
                valid = True
                return policy
            elif int(policyOption) == 100:
                policy = 'create_policy'
                valid = True
                return policy
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Selection.  Please Select a valid Index from the List.')
            print(f'\n-------------------------------------------------------------------------------------------\n')

def policies_parse(org, policy_type, policy):
    if os.environ.get('TF_DEST_DIR') is None:
        tfDir = 'Intersight'
    else:
        tfDir = os.environ.get('TF_DEST_DIR')
    policies = []

    opSystem = platform.system()
    if opSystem == 'Windows':
        policy_file = f'.\{tfDir}\{org}\{policy_type}\{policy}.auto.tfvars'
    else:
        policy_file = f'./{tfDir}/{org}/{policy_type}/{policy}.auto.tfvars'
    if os.path.isfile(policy_file):
        if len(policy_file) > 0:
            if opSystem == 'Windows':
                cmd = 'hcl2json.exe %s' % (policy_file)
            else:
                cmd = 'hcl2json %s' % (policy_file)
                # cmd = 'json2hcl -reverse < %s' % (policy_file)
            p = subprocess.run(
                cmd,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT
            )
            if 'unable to parse' in p.stdout.decode('utf-8'):
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  !!!! Encountered Error in Attempting to read file !!!!')
                print(f'  - {policy_file}')
                print(f'  Error was:')
                print(f'  - {p.stdout.decode("utf-8")}')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                json_data = {}
                return policies,json_data
            else:
                json_data = json.loads(p.stdout.decode('utf-8'))
                for i in json_data[policy]:
                    policies.append(i)
                return policies,json_data
    else:
        json_data = {}
        return policies,json_data

def policy_descr(name, policy_type):
    valid = False
    while valid == False:
        descr = input(f'What is the Description for the {policy_type}?  [{name} {policy_type}]: ')
        if descr == '':
            descr = '%s %s' % (name, policy_type)
        valid = validating.description(f'{policy_type} templateVars["descr"]', descr, 1, 62)
        if valid == True:
            return descr

def policy_name(namex, policy_type):
    valid = False
    while valid == False:
        name = input(f'What is the Name for the {policy_type}?  [{namex}]: ')
        if name == '':
            name = '%s' % (namex)
        valid = validating.name_rule(f'{policy_type} Name', name, 1, 62)
        if valid == True:
            return name

# Function to validate input for each method
def process_kwargs(required_args, optional_args, **kwargs):
    # Validate all required kwargs passed
    # if all(item in kwargs for item in required_args.keys()) is not True:
    #    error_ = '\n***ERROR***\nREQUIRED Argument Not Found in Input:\n "%s"\nInsufficient required arguments.' % (item)
    #    raise InsufficientArgs(error_)
    error_count = 0
    error_list = []
    for item in required_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs.keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following Optional Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]

    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]
    # Combine option and required dicts for Jinja template render
    templateVars = {**required_args, **optional_args}
    return(templateVars)

def process_method(wr_method, dest_dir, dest_file, template, **templateVars):
    opSystem = platform.system()
    if opSystem == 'Windows':
        if os.environ.get('TF_DEST_DIR') is None:
            tfDir = 'Intersight'
        else:
            tfDir = os.environ.get('TF_DEST_DIR')
        if re.search(r'^\\.*\\$', tfDir):
            dest_dir = '%s%s\%s' % (tfDir, templateVars["org"], dest_dir)
        elif re.search(r'^\\.*\w', tfDir):
            dest_dir = '%s\%s\%s' % (tfDir, templateVars["org"], dest_dir)
        else:
            dest_dir = '.\%s\%s\%s' % (tfDir, templateVars["org"], dest_dir)
        if not os.path.isdir(dest_dir):
            mk_dir = 'mkdir %s' % (dest_dir)
            os.system(mk_dir)
        dest_file_path = '%s\%s' % (dest_dir, dest_file)
        if not os.path.isfile(dest_file_path):
            create_file = 'type nul >> %s' % (dest_file_path)
            os.system(create_file)
        tf_file = dest_file_path
        wr_file = open(tf_file, wr_method)
    else:
        if os.environ.get('TF_DEST_DIR') is None:
            tfDir = 'Intersight'
        else:
            tfDir = os.environ.get('TF_DEST_DIR')
        if re.search(r'^\/.*\/$', tfDir):
            dest_dir = '%s%s/%s' % (tfDir, templateVars["org"], dest_dir)
        elif re.search(r'^\/.*\w', tfDir):
            dest_dir = '%s/%s/%s' % (tfDir, templateVars["org"], dest_dir)
        else:
            dest_dir = './%s/%s/%s' % (tfDir, templateVars["org"], dest_dir)
        if not os.path.isdir(dest_dir):
            mk_dir = 'mkdir -p %s' % (dest_dir)
            os.system(mk_dir)
        dest_file_path = '%s/%s' % (dest_dir, dest_file)
        if not os.path.isfile(dest_file_path):
            create_file = 'touch %s' % (dest_file_path)
            os.system(create_file)
        tf_file = dest_file_path
        wr_file = open(tf_file, wr_method)

    # Render Payload and Write to File
    payload = template.render(templateVars)
    wr_file.write(payload)
    wr_file.close()

# Function to Read Excel Workbook Data
def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

def sensitive_var_value(jsonData, **templateVars):
    sensitive_var = 'TF_VAR_%s' % (templateVars['Variable'])
    # -------------------------------------------------------------------------------------------------------------------------
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #--------------------------------------------------------------------------------------------------------------------------
    if os.environ.get(sensitive_var) is None:
        print(f"\n----------------------------------------------------------------------------------\n")
        print(f"  The Script did not find {sensitive_var} as an 'environment' variable.")
        print(f"  To not be prompted for the value of {templateVars['Variable']} each time")
        print(f"  add the following to your local environemnt:\n")
        print(f"    - Linux: export {sensitive_var}='{templateVars['Variable']}_value'")
        print(f"    - Windows: $env:{sensitive_var}='{templateVars['Variable']}_value'")
        print(f"\n----------------------------------------------------------------------------------\n")

    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '':
                valid = True

        valid = False
        while valid == False:
            if templateVars.get('Multi_Line_Input'):
                print(f'Enter the value for {templateVars["Variable"]}:')
                lines = []
                while True:
                    # line = input('')
                    line = stdiomask.getpass(prompt='')
                    if line:
                        lines.append(line)
                    else:
                        break
                if not re.search('(certificate|private_key)', sensitive_var):
                    secure_value = '\\n'.join(lines)
                else:
                    secure_value = '\n'.join(lines)
            else:
                valid_pass = False
                while valid_pass == False:
                    password1 = stdiomask.getpass(prompt=f'Enter the value for {templateVars["Variable"]}: ')
                    password2 = stdiomask.getpass(prompt=f'Re-Enter the value for {templateVars["Variable"]}: ')
                    if password1 == password2:
                        secure_value = password1
                        valid_pass = True
                    else:
                        print('!!!Error!!! Sensitive Values did not match.  Please re-enter...')


            # Validate Sensitive Passwords
            cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            if re.search('(certificate|private_key)', sensitive_var):
                if not re.search(cert_regex, secure_value):
                    valid = True
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'    Error!!! Invalid Value for the {sensitive_var}.  Please re-enter the {sensitive_var}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            elif re.search('(apikey|secretkey)', sensitive_var):
                if not sensitive_var == '':
                    valid = True
            elif 'bind' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['iam.LdapBaseProperties']['allOf'][1]['properties']
                minLength = 1
                maxLength = 254
                rePattern = jsonVars['Password']['pattern']
                varName = 'SNMP Community'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'community' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['snmp.Policy']['allOf'][1]['properties']
                minLength = 1
                maxLength = jsonVars['TrapCommunity']['maxLength']
                rePattern = '^[\\S]+$'
                varName = 'SNMP Community'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'ipmi_key' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['ipmioverlan.Policy']['allOf'][1]['properties']
                minLength = 2
                maxLength = jsonVars['EncryptionKey']['maxLength']
                rePattern = jsonVars['EncryptionKey']['pattern']
                varName = 'IPMI Encryption Key'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'iscsi_boot' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['vnic.IscsiAuthProfile']['allOf'][1]['properties']
                minLength = 12
                maxLength = 16
                rePattern = jsonVars['Password']['pattern']
                varName = 'iSCSI Boot Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'local' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['iam.EndPointUserRole']['allOf'][1]['properties']
                minLength = jsonVars['Password']['minLength']
                maxLength = jsonVars['Password']['maxLength']
                rePattern = jsonVars['Password']['pattern']
                varName = 'Local User Password'
                if templateVars.get('enforce_strong_password'):
                    enforce_pass = templateVars['enforce_strong_password']
                else:
                    enforce_pass = False
                if enforce_pass == True:
                    minLength = 8
                    maxLength = 20
                    valid = validating.strong_password(templateVars['Variable'], secure_value, minLength, maxLength)
                else:
                    valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'secure_passphrase' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['memory.PersistentMemoryLocalSecurity']['allOf'][1]['properties']
                minLength = jsonVars['SecurePassphrase']['minLength']
                maxLength = jsonVars['SecurePassphrase']['maxLength']
                rePattern = jsonVars['SecurePassphrase']['pattern']
                varName = 'Persistent Memory Secure Passphrase'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'snmp' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['snmp.Policy']['allOf'][1]['properties']
                minLength = 1
                maxLength = jsonVars['TrapCommunity']['maxLength']
                rePattern = '^[\\S]+$'
                if 'auth' in sensitive_var:
                    varName = 'SNMP Authorization Password'
                else:
                    varName = 'SNMP Privacy Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'vmedia' in sensitive_var:
                jsonVars = jsonData['components']['schemas']['vmedia.Mapping']['allOf'][1]['properties']
                minLength = 1
                maxLength = jsonVars['Password']['maxLength']
                rePattern = '^[\\S]+$'
                varName = 'vMedia Mapping Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)
        var_value = secure_value

    else:
        # Add the Variable to the Environment
        if templateVars.get('Multi_Line_Input'):
            var_value = os.environ.get(sensitive_var)
            var_value = var_value.replace('\n', '\\n')
        else:
            var_value = os.environ.get(sensitive_var)

    return var_value

def snmp_trap_servers(jsonData, inner_loop_count, snmp_user_list, **templateVars):
    trap_servers = []
    valid_traps = False
    while valid_traps == False:
        templateVars["multi_select"] = False
        jsonVars = jsonData['components']['schemas']['snmp.Trap']['allOf'][1]['properties']
        if len(snmp_user_list) == 0:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  There are no valid SNMP Users so Trap Destinations can only be set to SNMPv2.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            snmp_version = 'V2'
        else:
            templateVars["var_description"] = jsonVars['Version']['description']
            templateVars["jsonVars"] = sorted(jsonVars['Version']['enum'])
            templateVars["defaultVar"] = jsonVars['Version']['default']
            templateVars["varType"] = 'SNMP Version'
            snmp_version = variablesFromAPI(**templateVars)

        if snmp_version == 'V2':
            valid = False
            while valid == False:
                community_string = stdiomask.getpass(f'What is the Community String for the Destination? ')
                if not community_string == '':
                    valid = validating.snmp_string('SNMP Community String', community_string)
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Error!! Invalid Value.  Please Re-enter the SNMP Community String.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            TF_VAR = 'TF_VAR_snmp_community_string_%s' % (inner_loop_count)
            os.environ[TF_VAR] = '%s' % (community_string)
            community_string = inner_loop_count

        if snmp_version == 'V3':
            templateVars["multi_select"] = False
            templateVars["var_description"] = '    Please Select the SNMP User to assign to this Destination:\n'
            templateVars["var_type"] = 'SNMP User'
            snmp_users = []
            for item in snmp_user_list:
                snmp_users.append(item['name'])
            snmp_user = vars_from_list(snmp_users, **templateVars)
            snmp_user = snmp_user[0]

        if snmp_version == 'V2':
            templateVars["var_description"] = jsonVars['Type']['description']
            templateVars["jsonVars"] = sorted(jsonVars['Type']['enum'])
            templateVars["defaultVar"] = jsonVars['Type']['default']
            templateVars["varType"] = 'SNMP Trap Type'
            trap_type = variablesFromAPI(**templateVars)
        else:
            trap_type = 'Trap'

        valid = False
        while valid == False:
            destination_address = input(f'What is the SNMP Trap Destination Hostname/Address? ')
            if not destination_address == '':
                if re.search(r'^[0-9a-fA-F]+[:]+[0-9a-fA-F]$', destination_address) or \
                    re.search(r'^(\d{1,3}\.){3}\d{1,3}$', destination_address):
                    valid = validating.ip_address('SNMP Trap Destination', destination_address)
                else:
                    valid = validating.dns_name('SNMP Trap Destination', destination_address)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please Re-enter the SNMP Trap Destination Hostname/Address.')
                print(f'\n-------------------------------------------------------------------------------------------\n')

        valid = False
        while valid == False:
            port = input(f'Enter the Port to Assign to this Destination.  Valid Range is 1-65535.  [162]: ')
            if port == '':
                port = 162
            if re.search(r'[0-9]{1,4}', str(port)):
                valid = validating.snmp_port('SNMP Port', port, 1, 65535)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Invalid Entry!  Please Enter a valid Port in the range of 1-65535.')
                print(f'\n-------------------------------------------------------------------------------------------\n')

        if snmp_version == 'V3':
            snmp_destination = {
                'destination_address':destination_address,
                'enabled':True,
                'port':port,
                'trap_type':trap_type,
                'user':snmp_user,
                'version':snmp_version
            }
        else:
            snmp_destination = {
                'community':community_string,
                'destination_address':destination_address,
                'enabled':True,
                'port':port,
                'trap_type':trap_type,
                'version':snmp_version
            }

        print(f'\n-------------------------------------------------------------------------------------------\n')
        if snmp_version == 'V2':
            print(f'   community_string    = "Sensitive"')
        print(f'   destination_address = "{destination_address}"')
        print(f'   enable              = True')
        print(f'   trap_type           = "{trap_type}"')
        print(f'   snmp_version        = "{snmp_version}"')
        if snmp_version == 'V3':
            print(f'   user                = "{snmp_user}"')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_v = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if confirm_v == 'Y' or confirm_v == '':
                trap_servers.append(snmp_destination)
                valid_exit = False
                while valid_exit == False:
                    loop_exit = input(f'Would You like to Configure another SNMP Trap Destination?  Enter "Y" or "N" [N]: ')
                    if loop_exit == 'Y':
                        inner_loop_count += 1
                        valid_confirm = True
                        valid_exit = True
                    elif loop_exit == 'N' or loop_exit == '':
                        snmp_loop = True
                        valid_confirm = True
                        valid_exit = True
                        valid_traps = True
                    else:
                        print(f'\n------------------------------------------------------\n')
                        print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                        print(f'\n------------------------------------------------------\n')

            elif confirm_v == 'N':
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Starting Remote Host Configuration Over.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                valid_confirm = True
            else:
                print(f'\n------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                print(f'\n------------------------------------------------------\n')

    return trap_servers,snmp_loop

def snmp_users(jsonData, inner_loop_count, **templateVars):
    snmp_user_list = []
    valid_users = False
    while valid_users == False:
        templateVars["multi_select"] = False
        jsonVars = jsonData['components']['schemas']['snmp.User']['allOf'][1]['properties']

        snmpUser = False
        while snmpUser == False:
            templateVars["Description"] = jsonVars['Name']['description']
            templateVars["varDefault"] = 'admin'
            templateVars["varInput"] = 'What is the SNMPv3 Username:'
            templateVars["varName"] = 'SNMP User'
            templateVars["varRegex"] = '^([a-zA-Z]+[a-zA-Z0-9\\-\\_\\.\\@]+)$'
            templateVars["minLength"] = jsonVars['Name']['minLength']
            templateVars["maxLength"] = jsonVars['Name']['maxLength']
            snmp_user = varStringLoop(**templateVars)
            if snmp_user == 'admin':
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  admin may not be used for the snmp user value.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
            else:
                snmpUser = True

        templateVars["var_description"] = jsonVars['SecurityLevel']['description']
        templateVars["jsonVars"] = sorted(jsonVars['SecurityLevel']['enum'])
        templateVars["defaultVar"] = jsonVars['SecurityLevel']['default']
        templateVars["varType"] = 'SNMP Security Level'
        security_level = variablesFromAPI(**templateVars)

        if security_level == 'AuthNoPriv' or security_level == 'AuthPriv':
            templateVars["var_description"] = jsonVars['AuthType']['description']
            templateVars["jsonVars"] = sorted(jsonVars['AuthType']['enum'])
            templateVars["defaultVar"] = 'SHA'
            templateVars["popList"] = ['NA', 'SHA-224', 'SHA-256', 'SHA-384', 'SHA-512']
            templateVars["varType"] = 'SNMP Auth Type'
            auth_type = variablesFromAPI(**templateVars)

        if security_level == 'AuthNoPriv' or security_level == 'AuthPriv':
            valid = False
            while valid == False:
                password1 = stdiomask.getpass(f'What is the authorization password for {snmp_user}? ')
                password2 = stdiomask.getpass(f'Please re-enter the authorization password for {snmp_user}? ')
                if not password1 == '':
                    if password1 == password2:
                        TF_VAR = 'TF_VAR_snmp_auth_password_%s' % (inner_loop_count)
                        os.environ[TF_VAR] = '%s' % (password1)
                        auth_password = inner_loop_count
                        valid = validating.snmp_string(f"{snmp_user}'s Authorization Password", password1)
                    else:
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                        print(f'  Error!! The Passwords did not match.  Please Re-enter the password for {snmp_user}.')
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Error!! Invalid Value.  Please Re-enter the password for {snmp_user}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')

        if security_level == 'AuthPriv':
            templateVars["var_description"] = jsonVars['PrivacyType']['description']
            templateVars["jsonVars"] = sorted(jsonVars['PrivacyType']['enum'])
            templateVars["defaultVar"] = 'AES'
            templateVars["popList"] = ['NA']
            templateVars["varType"] = 'SNMP Auth Type'
            privacy_type = variablesFromAPI(**templateVars)

            valid = False
            while valid == False:
                password1 = stdiomask.getpass(f'What is the privacy password for {snmp_user}? ')
                password2 = stdiomask.getpass(f'Please re-enter the privacy password for {snmp_user}? ')
                if not password1 == '':
                    if password1 == password2:
                        TF_VAR = 'TF_VAR_snmp_privacy_password_%s' % (inner_loop_count)
                        os.environ[TF_VAR] = '%s' % (password1)
                        privacy_password = inner_loop_count
                        valid = validating.snmp_string(f"{snmp_user}'s Privacy Password", password1)
                    else:
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                        print(f'  Error!! The Passwords did not match.  Please Re-enter the password for {snmp_user}.')
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Error!! Invalid Value.  Please Re-enter the password for {snmp_user}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')

        if security_level == 'AuthPriv':
            snmp_user = {
                'auth_password':auth_password,
                'auth_type':auth_type,
                'name':snmp_user,
                'privacy_password':privacy_password,
                'privacy_type':privacy_type,
                'security_level':security_level
            }
        elif security_level == 'AuthNoPriv':
            snmp_user = {
                'auth_password':auth_password,
                'auth_type':auth_type,
                'name':snmp_user,
                'security_level':security_level
            }

        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'   auth_password    = "Sensitive"')
        print(f'   auth_type        = "{auth_type}"')
        if security_level == 'AuthPriv':
            print(f'   privacy_password = "Sensitive"')
            print(f'   privacy_type     = "{privacy_type}"')
        print(f'   security_level   = "{security_level}"')
        print(f'   snmp_user        = "{snmp_user["name"]}"')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_v = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if confirm_v == 'Y' or confirm_v == '':
                snmp_user_list.append(snmp_user)
                valid_exit = False
                while valid_exit == False:
                    loop_exit = input(f'Would You like to Configure another SNMP User?  Enter "Y" or "N" [N]: ')
                    if loop_exit == 'Y':
                        inner_loop_count += 1
                        valid_confirm = True
                        valid_exit = True
                    elif loop_exit == 'N' or loop_exit == '':
                        snmp_loop = True
                        valid_confirm = True
                        valid_exit = True
                        valid_users = True
                    else:
                        print(f'\n------------------------------------------------------\n')
                        print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                        print(f'\n------------------------------------------------------\n')

            elif confirm_v == 'N':
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Starting SNMP User Configuration Over.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                valid_confirm = True
            else:
                print(f'\n------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                print(f'\n------------------------------------------------------\n')
    return snmp_user_list,snmp_loop

# Function to Define stdout_log output
def stdout_log(sheet, line):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (sheet) and (line is None)):
        #print('*' * 80)
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Starting work on {sheet.title} Worksheet')
        print(f'\n-----------------------------------------------------------------------------\n')
        #print('*' * 80)
    elif log_level == (2) and (sheet) and (line is not None):
        print('Evaluating line %s from %s Worksheet...' % (line, sheet.title))
    else:
        return

def syslog_servers(jsonData, **templateVars):
    remote_logging = {}
    syslog_count = 1
    syslog_loop = False
    while syslog_loop == False:
        valid = False
        while valid == False:
            hostname = input(f'Enter the Hostname/IP Address of the Remote Server: ')
            if re.search(r'[a-zA-Z]+', hostname):
                valid = validating.dns_name('Remote Logging Server', hostname)
            else:
                valid = validating.ip_address('Remote Logging Server', hostname)

        jsonVars = jsonData['components']['schemas']['syslog.RemoteClientBase']['allOf'][1]['properties']
        templateVars["var_description"] = jsonVars['MinSeverity']['description']
        templateVars["jsonVars"] = sorted(jsonVars['MinSeverity']['enum'])
        templateVars["defaultVar"] = jsonVars['MinSeverity']['default']
        templateVars["varType"] = 'Syslog Remote Minimum Severity'
        min_severity = variablesFromAPI(**templateVars)

        templateVars["var_description"] = jsonVars['Protocol']['description']
        templateVars["jsonVars"] = sorted(jsonVars['Protocol']['enum'])
        templateVars["defaultVar"] = jsonVars['Protocol']['default']
        templateVars["varType"] = 'Syslog Protocol'
        templateVars["protocol"] = variablesFromAPI(**templateVars)

        valid = False
        while valid == False:
            port = input(f'Enter the Port to Assign to this Policy.  Valid Range is 1-65535.  [514]: ')
            if port == '':
                port = 514
            if re.search(r'[0-9]{1,4}', str(port)):
                valid = validating.number_in_range('Port', port, 1, 65535)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Invalid Entry!  Please Enter a valid Port in the range of 1-65535.')
                print(f'\n-------------------------------------------------------------------------------------------\n')

        remote_host = {
            'enable':True,
            'hostname':hostname,
            'min_severity':min_severity,
            'port':port,
            'protocol':templateVars["protocol"]
        }
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'   hostname     = "{hostname}"')
        print(f'   min_severity = "{min_severity}"')
        print(f'   port         = {port}')
        print(f'   protocol     = "{templateVars["protocol"]}"')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_host = input('Do you want to accept the configuration above?  Enter "Y" or "N" [Y]: ')
            if confirm_host == 'Y' or confirm_host == '':
                if syslog_count == 1:
                    remote_logging.update({'server1':remote_host})
                elif syslog_count == 2:
                    remote_logging.update({'server2':remote_host})
                    syslog_loop = True
                    valid_confirm = True
                if syslog_count == 1:
                    valid_exit = False
                    while valid_exit == False:
                        remote_exit = input(f'Would You like to Configure another Remote Host?  Enter "Y" or "N" [Y]: ')
                        if remote_exit == 'Y' or remote_exit == '':
                            syslog_count += 1
                            valid_confirm = True
                            valid_exit = True
                        elif remote_exit == 'N':
                            remote_host = {
                                'enable':False,
                                'hostname':'0.0.0.0',
                                'min_severity':'warning',
                                'port':514,
                                'protocol':'udp'
                            }
                            remote_logging.update({'server2':remote_host})
                            syslog_loop = True
                            valid_confirm = True
                            valid_exit = True
                        else:
                            print(f'\n------------------------------------------------------\n')
                            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                            print(f'\n------------------------------------------------------\n')

            elif confirm_host == 'N':
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Starting Syslog Server Configuration Over.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                valid_confirm = True
            else:
                print(f'\n------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                print(f'\n------------------------------------------------------\n')

    return remote_logging

#======================================================
# Function to Format Terraform Files
#======================================================
def terraform_fmt(files, folder, path_sep):
    # Create the Empty_variable_maps.auto.tfvars to house all the unused variables
    empty_auto_tfvars = f'{folder}{path_sep}Empty_variable_maps.auto.tfvars'
    wr_file = open(empty_auto_tfvars, 'w')
    wrString = f'#______________________________________________'\
              '\n#'\
              '\n# UNUSED Variables'\
              '\n#______________________________________________\n\n'
    wr_file.write(wrString)
    for file in files:
        varFiles = f"{file.split('.')[0]}.auto.tfvars"
        dest_file = f'{folder}{path_sep}{varFiles}'
        if not os.path.isfile(dest_file):
            x = file.split('.')
            if re.search('(ndo_sites|ndo_users)', x[0]):
                wrString = f'{x[0]} = ''[]\n'
            else:
                wrString = f'{x[0]} = ''{}\n'
            wr_file.write(wrString)

    # Close the Unused Variables File
    wr_file.close()

    # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  Running "terraform fmt" in folder "{folder}",')
    print(f'  to correct variable formatting!')
    print(f'\n-------------------------------------------------------------------------------------------\n')
    p = subprocess.Popen(
        ['terraform', 'fmt', folder],
        stdout = subprocess.PIPE,
        stderr = subprocess.PIPE
    )
    print('Format updated for the following Files:')
    for line in iter(p.stdout.readline, b''):
        line = line.decode("utf-8")
        line = line.strip()
        print(f'- {line}')

def tfc_sensitive_variables(varValue, jsonData, templateVars):
    templateVars["Variable"] = varValue
    if 'ipmi_key' in varValue:
        templateVars["Description"] = 'IPMI over LAN Encryption Key'
    elif 'iscsi' in varValue:
        templateVars["Description"] = 'iSCSI Boot Password'
    elif 'local_user' in varValue:
        templateVars["Description"] = 'Local User Password'
    elif 'access_comm' in varValue:
        templateVars["Description"] = 'SNMP Access Community String'
    elif 'snmp_auth' in varValue:
        templateVars["Description"] = 'SNMP Authorization Password'
    elif 'snmp_priv' in varValue:
        templateVars["Description"] = 'SNMP Privacy Password'
    elif 'trap_comm' in varValue:
        templateVars["Description"] = 'SNMP Trap Community String'
    templateVars["varValue"] = sensitive_var_value(jsonData, **templateVars)
    templateVars["varId"] = varValue
    templateVars["varKey"] = varValue
    templateVars["Sensitive"] = True
    print(f'* Adding {templateVars["Description"]} to {templateVars["workspaceName"]}')
    return templateVars

def ucs_domain_serials():
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  Note: If you do not have the Serial Number at this time you can manually add it to the:')
    print(f'        - ucs_domain_profiles/ucs_domain_profiles.auto.tfvars file later.')
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        templateVars = {}
        fabrics = ['A','B']
        for x in fabrics:
            templateVars[f"serial_{x}"] = input(f'What is the Serial Number of Fabric {x}? [press enter to skip]: ')
            if templateVars[f"serial_{x}"] == '':
                valid = True
            elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', templateVars[f"serial_{x}"]):
                valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Serial Number.  "templateVars["serial_{x}"]" is not a valid serial.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
    serial_a = templateVars["serial_A"]
    serial_b = templateVars["serial_B"]
    return serial_a,serial_b

def validate_vlan_in_policy(vlan_policy_list, vlan_id):
    valid = False
    vlan_count = 0
    for vlan in vlan_policy_list:
        if int(vlan) == int(vlan_id):
            vlan_count = 1
            continue
    if vlan_count == 1:
        valid = True
    else:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'  VLAN {vlan_id} not found in the VLAN Policy List.  Please us a VLAN from the list below:')
        print(f'  {vlan_policy_list}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
    return valid

def variablesFromAPI(**templateVars):
    valid = False
    while valid == False:
        json_vars = templateVars["jsonVars"]
        if 'popList' in templateVars:
            if len(templateVars["popList"]) > 0:
                for x in templateVars["popList"]:
                    varsCount = len(json_vars)
                    for r in range(0, varsCount):
                        if json_vars[r] == x:
                            json_vars.pop(r)
                            break
        print(f'\n-------------------------------------------------------------------------------------------\n')
        newDescr = templateVars["var_description"]
        if '\n' in newDescr:
            newDescr = newDescr.split('\n')
            for line in newDescr:
                if '*' in line:
                    print(fill(f'{line}',width=88, subsequent_indent='    '))
                else:
                    print(fill(f'{line}',88))
        else:
            print(fill(f'{templateVars["var_description"]}',88))
        print(f'\n    Select an Option Below:')
        for index, value in enumerate(json_vars):
            index += 1
            if value == templateVars["defaultVar"]:
                defaultIndex = index
            if index < 10:
                print(f'     {index}. {value}')
            else:
                print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if templateVars["multi_select"] == True:
            if not templateVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number(s) to Select for {templateVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number(s) to Select for {templateVars["varType"]}: ')
        else:
            if not templateVars["defaultVar"] == '':
                var_selection = input(f'Please Enter the Option Number to Select for {templateVars["varType"]}.  [{defaultIndex}]: ')
            else:
                var_selection = input(f'Please Enter the Option Number to Select for {templateVars["varType"]}: ')
        if not templateVars["defaultVar"] == '' and var_selection == '':
            var_selection = defaultIndex

        if templateVars["multi_select"] == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(json_vars):
                index += 1
                if int(var_selection) == index:
                    selection = value
                    valid = True
        elif templateVars["multi_select"] == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(json_vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index:
                        var_count += 1
                        selection.append(value)
            if var_count == var_length:
                valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  The list of Vars {var_list} did not match the available list.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return selection

def varBoolLoop(**templateVars):
    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]}  [{templateVars["varDefault"]}]: ')
        if varValue == '':
            if templateVars["varDefault"] == 'Y':
                varValue = True
            elif templateVars["varDefault"] == 'N':
                varValue = False
            valid = True
        elif varValue == 'N':
            varValue = False
            valid = True
        elif varValue == 'Y':
            varValue = True
            valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {templateVars["varName"]} value of "{varValue}" is Invalid!!! Please enter "Y" or "N".')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

def varNumberLoop(**templateVars):
    maxNum = templateVars["maxNum"]
    minNum = templateVars["minNum"]
    varName = templateVars["varName"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]}  [{templateVars["varDefault"]}]: ')
        if varValue == '':
            varValue = templateVars["varDefault"]
        if re.fullmatch(r'^[0-9]+$', str(varValue)):
            valid = validating.number_in_range(varName, varValue, minNum, maxNum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'   Valid range is {minNum} to {maxNum}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

def varSensitiveStringLoop(**templateVars):
    maxLength = templateVars["maxLength"]
    minLength = templateVars["minLength"]
    varName = templateVars["varName"]
    varRegex = templateVars["varRegex"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = stdiomask.getpass(f'{templateVars["varInput"]} ')
        if not varValue == '':
            valid = validating.length_and_regex_sensitive(varRegex, varName, varValue, minLength, maxLength)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

def varStringLoop(**templateVars):
    maxLength = templateVars["maxLength"]
    minLength = templateVars["minLength"]
    varName = templateVars["varName"]
    varRegex = templateVars["varRegex"]

    print(f'\n-------------------------------------------------------------------------------------------\n')
    newDescr = templateVars["Description"]
    if '\n' in newDescr:
        newDescr = newDescr.split('\n')
        for line in newDescr:
            if '*' in line:
                print(fill(f'{line}',width=88, subsequent_indent='    '))
            else:
                print(fill(f'{line}',88))
    else:
        print(fill(f'{templateVars["Description"]}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{templateVars["varInput"]} ')
        if 'press enter to skip' in templateVars["varInput"] and varValue == '':
            valid = True
        elif not templateVars["varDefault"] == '' and varValue == '':
            varValue = templateVars["varDefault"]
            valid = True
        elif not varValue == '':
            valid = validating.length_and_regex(varRegex, varName, varValue, minLength, maxLength)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

def vars_from_list(var_options, **templateVars):
    selection = []
    selection_count = 0
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'{templateVars["var_description"]}')
        for index, value in enumerate(var_options):
            index += 1
            if index < 10:
                print(f'     {index}. {value}')
            else:
                print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        exit_answer = False
        while exit_answer == False:
            var_selection = input(f'Please Enter the Option Number to Select for {templateVars["var_type"]}: ')
            if not var_selection == '':
                if re.search(r'[0-9]+', str(var_selection)):
                    xcount = 1
                    for index, value in enumerate(var_options):
                        index += 1
                        if int(var_selection) == index:
                            selection.append(value)
                            xcount = 0
                    if xcount == 0:
                        if selection_count % 2 == 0 and templateVars["multi_select"] == True:
                            answer_finished = input(f'Would you like to add another port to the {templateVars["port_type"]}?  Enter "Y" or "N" [Y]: ')
                        elif templateVars["multi_select"] == True:
                            answer_finished = input(f'Would you like to add another port to the {templateVars["port_type"]}?  Enter "Y" or "N" [N]: ')
                        elif templateVars["multi_select"] == False:
                            answer_finished = 'N'
                        if (selection_count % 2 == 0 and answer_finished == '') or answer_finished == 'Y':
                            exit_answer = True
                            selection_count += 1
                        elif answer_finished == '' or answer_finished == 'N':
                            exit_answer = True
                            valid = True
                        elif templateVars["multi_select"] == False:
                            exit_answer = True
                            valid = True
                        else:
                            print(f'\n------------------------------------------------------\n')
                            print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
                            print(f'\n------------------------------------------------------\n')
                    else:
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                        print(f'  Error!! Invalid Selection.  Please select a valid option from the List.')
                        print(f'\n-------------------------------------------------------------------------------------------\n')

                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
    return selection

def vlan_list_format(vlan_list_expanded):
    vlanGroups = itertools.groupby(vlan_list_expanded, key=lambda item, c=itertools.count():item-next(c))
    tempvlans = [list(g) for k, g in vlanGroups]
    vlanList = [str(x[0]) if len(x) == 1 else "{}-{}".format(x[0],x[-1]) for x in tempvlans]
    vlan_list = ",".join(vlanList)
    return vlan_list

def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(int(vl))
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(int(v))
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

def vlan_pool():
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'  The allowed vlan list can be in the format of:')
        print(f'     5 - Single VLAN')
        print(f'     1-10 - Range of VLANs')
        print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
        print(f'     1-10,20-30 - Ranges and Lists of VLANs')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        VlanList = input('Enter the VLAN or List of VLANs to assign to the Domain VLAN Pool: ')
        if not VlanList == '':
            vlanListExpanded = vlan_list_full(VlanList)
            valid_vlan = True
            for vlan in vlanListExpanded:
                valid_vlan = validating.number_in_range('VLAN ID', vlan, 1, 4094)
                if valid_vlan == False:
                    continue
            if valid_vlan == False:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error with VLAN(s) assignment!!! VLAN List: "{VlanList}" is not Valid.')
                print(f'  The allowed vlan list can be in the format of:')
                print(f'     5 - Single VLAN')
                print(f'     1-10 - Range of VLANs')
                print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
                print(f'     1-10,20-30 - Ranges and Lists of VLANs')
                print(f'\n-------------------------------------------------------------------------------------------\n')
            else:
                valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  The allowed vlan list can be in the format of:')
            print(f'     5 - Single VLAN')
            print(f'     1-10 - Range of VLANs')
            print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
            print(f'     1-10,20-30 - Ranges and Lists of VLANs')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    
    return VlanList,vlanListExpanded

def write_to_template(self, **templateVars):
    # Define the Template Source
    template = self.templateEnv.get_template(templateVars["template_file"])

    # Process the template
    dest_dir = '%s' % (self.type)
    dest_file = '%s.auto.tfvars' % (templateVars["template_type"])
    if templateVars["initial_write"] == True:
        write_method = 'w'
    else:
        write_method = 'a'
    process_method(write_method, dest_dir, dest_file, template, **templateVars)
